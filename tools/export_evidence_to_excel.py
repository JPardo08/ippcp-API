#!/usr/bin/env python3
"""Export IPPCP JSON evidence runs to a readable Excel workbook."""

from __future__ import annotations

import argparse
import io
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from evidence_assets import (
    ClassificationError,
    classify_loader,
    is_minimal_profile,
    load_asset_registry,
    package_publication_status,
    slot_inventory_row,
    validate_classified_run,
)
from evidence_common import (
    CELL_PROTECTED_STRUCTURE,
    CELL_PROTECTED_TEMPLATE_TEXT,
    CELL_RUNTIME_POPULATED,
    CELL_SANITIZATION_ALLOWED,
    ConnectorSanitizer,
    EvidenceRunLoader,
    FileEntry,
    FileIndexer,
    MINIMAL_PUBLICATION_PROFILE,
    NOT_FOUND,
    SummaryParser,
    MinimalPublicationModel,
    WorkbookContract,
    audit_minimal_publication_workbook,
    audit_minimal_publication_xlsx_bytes,
    build_test_specs,
    extract_minimal_publication_model,
    find_repo_root,
    load_test_config,
    parse_only_tests,
    parse_tests_override,
    NO_SLOTS_SELECTED_ERROR,
    relative_to_repo,
    resolve_repo_path,
    workbook_cell_snapshot,
)


SUMMARY_COLUMNS = [
    "test_id",
    "asset_key",
    "display_name",
    "family",
    "variant",
    "transport",
    "critical",
    "publication_profile",
    "publication_safe",
    "workflow",
    "asset_type",
    "provider_connector",
    "consumer_connector",
    "suffix",
    "asset_id",
    "vocab_id",
    "access_policy_id",
    "contract_policy_id",
    "contract_definition_id",
    "offer_policy_id",
    "negotiation_id",
    "agreement_id",
    "transfer_id",
    "transfer_type",
    "transfer_state",
    "download_status",
    "download_file",
    "bytes",
    "sha256",
    "summary_json",
    "manifest_json",
    "overall_status",
    "notes",
]

SLOT_MAP_COLUMNS = [
    "slot",
    "asset",
    "asset_key",
    "asset_variant",
    "family",
    "variant",
    "transport",
    "critical",
    "publication_profile",
    "publication_safe",
    "suffix",
    "validation_status",
]

SANITIZATION_FIELD_LABELS = {
    "provider_connector",
    "consumer_connector",
    "technical_provider_connector",
    "technical_consumer_connector",
    "asset_config",
    "download_file",
    "summary_json",
    "manifest_json",
    "file_path",
    "evidence_file",
    "artifact",
}

TABLE_TITLES = {
    "Datos generales",
    "IDs principales",
    "Estado por fase",
    "Pasos (summary)",
    "Artefactos por fase",
    "Manifest",
    "Notas de interpretación",
    "Slot identity",
}

CHECKLIST_REQUIREMENTS = [
    "provider authentication",
    "consumer authentication",
    "asset created/published",
    "policy created",
    "contract definition created",
    "catalog discovery",
    "offer policy selected/validated",
    "contract negotiation started",
    "negotiation finalized / agreement obtained",
    "transfer started/completed",
    "data downloaded/fetched",
    "manifest/hash generated",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", help="YAML/JSON config with tests")
    parser.add_argument(
        "--tests",
        help="Exact slot mapping SLOT=SUFFIX,SLOT=SUFFIX. No other slots are added.",
    )
    parser.add_argument(
        "--preset",
        help="Named historical/canonical slot mapping from config (does not bind assets to slots)",
    )
    parser.add_argument(
        "--only-tests",
        help="Filter the selected slot set (--tests or --preset) to these IDs, comma-separated",
    )
    parser.add_argument("--repo-root", help="Repository root")
    parser.add_argument("--evidence-dir", help="Evidence runs directory")
    parser.add_argument("--downloads-dir", help="Downloads directory")
    parser.add_argument("--output", help="Output .xlsx path")
    parser.add_argument("--export-dir", help="Base directory for timestamped exports")
    parser.add_argument("--timestamp", help="Timestamp to use for generated names, format YYYYMMDD_HHMMSS")
    parser.add_argument("--timestamp-suffix", action="store_true", help="Append timestamp before .xlsx")
    parser.add_argument("--strict", action="store_true", help="Fail on warnings or incomplete tests")
    parser.add_argument("--profile", help="Policy profile from config (does not assign assets to slots)")
    parser.add_argument(
        "--sanitize-connectors",
        action=argparse.BooleanOptionalAction,
        default=None,
        help="Apply connector aliases in output (default: profile/policy)",
    )
    parser.add_argument(
        "--redact-local-paths",
        action=argparse.BooleanOptionalAction,
        default=None,
        help="Redact absolute repo paths in output (default: profile/policy)",
    )
    parser.add_argument("--verbose", action="store_true")
    return parser.parse_args()


def warn(message: str, warnings: List[str], verbose: bool = True) -> None:
    warnings.append(message)
    if verbose:
        print(f"WARN: {message}", file=sys.stderr)


def nested_id_from_file(path: Path) -> str:
    if not path.exists() or path.name.endswith(".sensitive.json"):
        return NOT_FOUND
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return NOT_FOUND
    return str(data.get("@id") or data.get("id") or NOT_FOUND)


def step_value(parser: SummaryParser, phase: str, step_id: str, key: str) -> str:
    step = parser.get_step(phase, step_id)
    if not step:
        return NOT_FOUND
    value = step.get(key)
    return str(value) if value not in (None, "") else NOT_FOUND


def first_step_value(parser: SummaryParser, candidates: List[Tuple[str, str, str]]) -> str:
    for phase, step_id, key in candidates:
        value = step_value(parser, phase, step_id, key)
        if value != NOT_FOUND:
            return value
    return NOT_FOUND


def manifest_data(path: Optional[Path]) -> Dict[str, Any]:
    if not path or not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def phase_is_ok(parser: SummaryParser, phase: str) -> bool:
    return parser.phase_status(phase) == "ok"


def compute_overall_status(spec, parser: SummaryParser, row: Dict[str, Any]) -> Tuple[str, str]:
    expected = spec.expected_phases or list(parser.phases.keys())
    missing = [phase for phase in expected if phase not in parser.phases]
    if missing:
        return "INCOMPLETE", f"Missing phases: {', '.join(missing)}"
    failed = [phase for phase in expected if parser.phase_status(phase) != "ok"]
    if failed:
        return "FAIL", f"Non-ok phases: {', '.join(failed)}"

    transfer_state = row.get("transfer_state")
    download_status = row.get("download_status")
    has_download = download_status == "ok" and str(row.get("sha256") or NOT_FOUND) != NOT_FOUND
    if not has_download:
        negotiation = parser.find_step("negotiation_finalized")
        if negotiation and negotiation[1].get("status") == "ok":
            return "PARTIAL", "Negotiation ok but download evidence missing or failed"
        return "INCOMPLETE", "Download evidence missing"

    if parser.detect_workflow_kind() == "b2" and transfer_state == "STARTED":
        return "PASS_WITH_NOTE", "B2 transfer is STARTED, but phase4b.storage_fetch verified consumer MinIO download with hash"
    if parser.detect_workflow_kind() == "b1" and transfer_state == "STARTED":
        return "PASS_WITH_NOTE", "B1 transfer is STARTED, but data_consumed and save_download verified download with hash"
    return "PASS", ""


def build_summary_row(loader: EvidenceRunLoader, spec, repo_root: Path) -> Dict[str, Any]:
    parser = loader.parser or SummaryParser(loader.summary)
    manifest_path = loader.canonical_manifest_path()
    manifest = manifest_data(manifest_path)
    kind = parser.detect_workflow_kind()
    publish_phase = "phase1b" if kind == "b2" else "phase1"
    transfer_phase = "phase3b" if kind == "b2" else "phase3"
    download_phase = "phase4b" if kind == "b2" else "phase4"
    download_step = "storage_fetch" if kind == "b2" else "save_download"

    row: Dict[str, Any] = {
        "test_id": spec.test_id,
        "asset_key": spec.asset_key,
        "display_name": spec.display_name,
        "family": spec.family,
        "variant": spec.variant,
        "transport": spec.transport,
        "critical": spec.critical,
        "publication_profile": spec.publication_profile,
        "publication_safe": spec.publication_safe,
        "workflow": spec.workflow,
        "asset_type": spec.asset_type,
        "provider_connector": spec.provider_connector,
        "consumer_connector": spec.consumer_connector,
        "suffix": loader.summary.get("suffix") or spec.suffix,
        "asset_id": loader.asset_id(),
        "vocab_id": loader.env.get("VOCAB_ID") or nested_id_from_file(loader.run_dir / publish_phase / "10_create_vocabulary.json"),
        "access_policy_id": loader.env.get("ACCESS_POLICY_ID") or nested_id_from_file(loader.run_dir / publish_phase / "11_create_access_policy.json"),
        "contract_policy_id": loader.env.get("CONTRACT_POLICY_ID") or nested_id_from_file(loader.run_dir / publish_phase / "12_create_contract_policy.json"),
        "contract_definition_id": loader.env.get("CD_ID") or first_existing_id(loader.run_dir / publish_phase, "create_contract_definition"),
        "offer_policy_id": first_step_value(parser, [("phase2", "selected_ids", "offer_policy_id")]),
        "negotiation_id": first_step_value(
            parser,
            [
                ("phase2", "contract_negotiation_started", "neg_id"),
                ("phase2", "negotiation_finalized", "neg_id"),
            ],
        ),
        "agreement_id": first_step_value(
            parser,
            [
                ("phase2", "negotiation_finalized", "agreement_id"),
                ("phase2", "get_contract_agreement", "agreement_id"),
            ],
        ),
        "transfer_id": first_step_value(
            parser,
            [
                (transfer_phase, "transfer_started", "transfer_id"),
                (transfer_phase, "transfer_final_state", "transfer_id"),
            ],
        ),
        "transfer_type": first_found(
            first_step_value(
                parser,
                [
                    (transfer_phase, "transfer_started", "transfer_type"),
                    (transfer_phase, "transfer_type_valid", "transfer_type"),
                    (transfer_phase, "transfer_final_state", "transfer_type"),
                ],
            ),
            manifest.get("transfer_type"),
        ),
        "transfer_state": first_found(
            first_step_value(parser, [(transfer_phase, "transfer_final_state", "final_state")]),
            manifest.get("consumer_transfer_state"),
        ),
        "download_status": first_step_value(parser, [(download_phase, download_step, "status")]),
        "download_file": first_found(
            first_step_value(
                parser,
                [
                    (download_phase, download_step, "latest_file"),
                    (download_phase, download_step, "download_file"),
                ],
            ),
            manifest.get("latest_file"),
        ),
        "bytes": first_found(first_step_value(parser, [(download_phase, download_step, "bytes")]), manifest.get("bytes")),
        "sha256": first_found(first_step_value(parser, [(download_phase, download_step, "sha256")]), manifest.get("sha256")),
        "summary_json": relative_to_repo(repo_root, loader.summary_path),
        "manifest_json": relative_to_repo(repo_root, manifest_path) if manifest_path and manifest_path.exists() else NOT_FOUND,
        "overall_status": "",
        "notes": "",
    }
    status, note = compute_overall_status(spec, parser, row)
    row["overall_status"] = status
    row["notes"] = note
    return {key: normalize_cell(row.get(key)) for key in SUMMARY_COLUMNS}


def first_existing_id(phase_dir: Path, name_fragment: str) -> str:
    if not phase_dir.exists():
        return NOT_FOUND
    for path in sorted(phase_dir.glob(f"*{name_fragment}*.json")):
        if path.name.endswith("_request.json") or path.name.endswith(".sensitive.json"):
            continue
        value = nested_id_from_file(path)
        if value != NOT_FOUND:
            return value
    return NOT_FOUND


def normalize_cell(value: Any) -> Any:
    if value is None or value == "":
        return NOT_FOUND
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def first_found(*values: Any) -> Any:
    for value in values:
        if value not in (None, "", NOT_FOUND):
            return value
    return NOT_FOUND


def minimal_publication_execution_outcome(model: MinimalPublicationModel) -> Tuple[str, str]:
    """Map canonical execution evidence to the compatible Summary outcome."""
    non_ok_phases = [
        phase for phase, status in model.phase_statuses.items() if status != "ok"
    ]
    if non_ok_phases:
        return "FAIL", f"Non-ok phases: {', '.join(non_ok_phases)}"
    if model.delivery_mode == "post_metadata_only":
        if model.download_status != model.not_applicable:
            return "FAIL", f"POST metadata-only download status: {model.download_status}"
        if model.http_operation != "POST" or model.http_method != "POST":
            return "FAIL", "POST metadata-only requires POST operation"
        try:
            code = int(model.http_status)
        except (TypeError, ValueError):
            return "FAIL", f"POST http_status invalid: {model.http_status}"
        if code < 200 or code >= 300:
            return "FAIL", f"POST http_status: {code}"
        return "PASS", "POST metadata-only"
    if model.download_status != "ok":
        return "FAIL", f"Download status: {model.download_status}"
    return "PASS", ""


def minimal_publication_summary_row(model: MinimalPublicationModel, spec=None) -> Dict[str, Any]:
    identifiers = model.execution_identifiers
    overall_status, outcome_note = minimal_publication_execution_outcome(model)
    row = {
        "test_id": model.test_id,
        "asset_key": spec.asset_key if spec else "",
        "display_name": model.public_flow_label,
        "family": spec.family if spec else "ingestion",
        "variant": spec.variant if spec else "api_v2",
        "transport": spec.transport if spec else "HttpData-PULL",
        "critical": spec.critical if spec else True,
        "publication_profile": (
            spec.publication_profile if spec else MINIMAL_PUBLICATION_PROFILE
        ),
        "publication_safe": spec.publication_safe if spec else True,
        "workflow": model.public_flow_label,
        "asset_type": model.asset_type,
        "provider_connector": model.technical_provider_connector,
        "consumer_connector": model.technical_consumer_connector,
        "suffix": identifiers["run_id"],
        "asset_id": identifiers["asset_id"],
        "vocab_id": model.not_applicable,
        "access_policy_id": model.not_applicable,
        "contract_policy_id": model.not_applicable,
        "contract_definition_id": identifiers["contract_definition_id"],
        "offer_policy_id": model.not_applicable,
        "negotiation_id": identifiers["negotiation_id"],
        "agreement_id": identifiers["agreement_id"],
        "transfer_id": identifiers["transfer_process_id"],
        "transfer_type": model.not_recorded,
        "transfer_state": model.not_recorded,
        "download_status": model.download_status,
        "download_file": model.not_applicable,
        "bytes": model.byte_count,
        "sha256": model.sha256_value,
        "summary_json": f"{(spec.sheet_name if spec else f'{model.test_id}_ingestion_api')}/sanitized_summary.json",
        "manifest_json": f"{(spec.sheet_name if spec else f'{model.test_id}_ingestion_api')}/sanitized_manifest.json",
        "overall_status": overall_status,
        "notes": (
            f"{model.evidence_role}; "
            f"delivery_mode={model.delivery_mode}; "
            f"semantic_validation={model.semantic_validation_status}"
            + (f"; {outcome_note}" if outcome_note else "")
        ),
    }
    return {column: row[column] for column in SUMMARY_COLUMNS}


def append_minimal_publication_sheet(
    workbook: Workbook,
    spec,
    model: MinimalPublicationModel,
    contract: Optional[WorkbookContract] = None,
) -> None:
    worksheet = workbook.create_sheet(spec.sheet_name[:31])
    append_table(
        worksheet,
        "Datos generales",
        ["field", "value"],
        [
            ["slot", spec.test_id],
            ["asset", spec.display_name or model.public_flow_label],
            ["asset_key", spec.asset_key],
            ["family", spec.family],
            ["variant", spec.variant],
            ["transport", spec.transport],
            ["critical", spec.critical],
            ["publication_profile", spec.publication_profile],
            ["publication_safe", spec.publication_safe],
            ["test_id", model.test_id],
            ["workflow", model.public_flow_label],
            ["asset_type", model.asset_type],
            ["evidence_role", model.evidence_role],
            ["technical_status", model.technical_status],
            ["technical_provider_connector", model.technical_provider_connector],
            ["technical_consumer_connector", model.technical_consumer_connector],
        ],
        contract=contract,
    )
    append_table(
        worksheet,
        "IDs principales",
        ["field", "value"],
        [[key, value] for key, value in model.execution_identifiers.items()]
        + [["sha256", model.sha256_value]],
        contract=contract,
    )
    append_table(
        worksheet,
        "Estado por fase",
        ["phase", "status", "step_count", "first_ts", "last_ts"],
        [
            [phase, status, model.not_recorded, model.not_recorded, model.not_recorded]
            for phase, status in model.phase_statuses.items()
        ],
        contract=contract,
    )
    append_table(
        worksheet,
        "Manifest",
        ["field", "value"],
        [
            ["delivery_mode", model.delivery_mode],
            ["operation", model.http_operation],
            ["http_method", model.http_method],
            ["http_status", model.http_status],
            ["manifest_kind", model.manifest_kind],
            ["request_body_persisted", model.request_body_persisted],
            ["response_body_persisted", model.response_body_persisted],
            ["download_persisted", model.download_persisted],
            ["download_status", model.download_status],
            ["byte_count", model.byte_count],
            ["sha256_algorithm", model.sha256_algorithm],
            ["sha256_verified", model.sha256_verified],
            ["sha256_value", model.sha256_value],
            ["payload_included", model.payload_included],
        ],
        contract=contract,
    )
    append_table(
        worksheet,
        "Notas de interpretación",
        ["field", "value"],
        [
            ["evidence_role", model.evidence_role],
            ["semantic_validation_status", model.semantic_validation_status],
            ["semantic_validation_source", model.semantic_validation_source],
        ],
        contract=contract,
    )


def minimal_publication_aggregate_rows(
    model: MinimalPublicationModel,
    folder: str,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    logical_files = [
        ("sanitized_summary.json", "summary"),
        ("sanitized_manifest.json", "manifest"),
        ("validation_status.json", "validation"),
    ]
    raw_rows = [
        {
            "test_id": model.test_id,
            "suffix": model.execution_identifiers["run_id"],
            "phase": model.not_applicable,
            "file_path": f"{folder}/{file_name}",
            "file_name": file_name,
            "is_sensitive": False,
            "include_in_package": True,
            "description": f"Sanitized publication {category}",
            "http_status_file": model.not_applicable,
            "related_step": model.not_applicable,
        }
        for file_name, category in logical_files
    ]
    phase_requirements = {
        "provider authentication": "phase0",
        "consumer authentication": "phase0",
        "asset created/published": "phase1",
        "policy created": "phase1",
        "contract definition created": "phase1",
        "catalog discovery": "phase2",
        "offer policy selected/validated": "phase2",
        "contract negotiation started": "phase2",
        "negotiation finalized / agreement obtained": "phase2",
        "transfer started/completed": "phase3",
        "data downloaded/fetched": "phase4",
        "manifest/hash generated": "phase4",
    }
    checklist = [
        {
            "test_id": model.test_id,
            "requirement": requirement,
            "status": model.phase_statuses[phase],
            "evidence_file": f"{folder}/sanitized_summary.json",
            "evidence_step": model.not_recorded,
            "notes": model.evidence_role,
        }
        for requirement, phase in phase_requirements.items()
    ]
    package = [
        {
            "test_id": model.test_id,
            "file_path": f"{folder}/{file_name}",
            "category": "publication_metadata",
            "include_in_package": True,
            "reason": "minimal_publication allowlist",
            "size_bytes": model.not_recorded,
        }
        for file_name, _ in logical_files
    ]
    return raw_rows, checklist, package


def sanitize_value(value: Any, config: Dict[str, Any], repo_root: Path, sanitize: bool, redact_paths: bool) -> Any:
    if not isinstance(value, str):
        return value
    aliases = config.get("connector_aliases") if sanitize else {}
    return ConnectorSanitizer.apply(value, aliases, redact_paths, repo_root)


def _record_contract_cell(
    contract: Optional[WorkbookContract],
    sheet: str,
    coordinate: str,
    category: str,
    value: Any,
    label: str = "",
) -> None:
    if contract is None or value in (None, ""):
        return
    contract.add(sheet, coordinate, category, value, label)


def append_table(
    ws,
    title: str,
    headers: List[str],
    rows: List[List[Any]],
    contract: Optional[WorkbookContract] = None,
) -> None:
    sheet = ws.title
    empty = ws.max_row == 1 and ws["A1"].value in (None, "")
    if empty:
        ws["A1"] = title
        title_row = 1
    else:
        ws.append([])
        ws.append([title])
        title_row = ws.max_row
    ws.cell(title_row, 1).font = Font(bold=True, size=12)
    _record_contract_cell(contract, sheet, f"A{title_row}", CELL_PROTECTED_STRUCTURE, title, title)
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        _record_contract_cell(
            contract,
            sheet,
            cell.coordinate,
            CELL_PROTECTED_TEMPLATE_TEXT,
            cell.value,
            str(cell.value or ""),
        )
    field_value_layout = headers[:2] == ["field", "value"]
    for row in rows:
        ws.append(row)
        data_row = ws.max_row
        label = str(row[0]) if row else ""
        for index, value in enumerate(row, 1):
            coordinate = f"{get_column_letter(index)}{data_row}"
            if index == 1 and field_value_layout:
                category = CELL_PROTECTED_TEMPLATE_TEXT
            elif field_value_layout and index == 2:
                category = (
                    CELL_SANITIZATION_ALLOWED
                    if label in SANITIZATION_FIELD_LABELS
                    else CELL_RUNTIME_POPULATED
                )
            else:
                header = headers[index - 1] if index <= len(headers) else ""
                category = (
                    CELL_SANITIZATION_ALLOWED
                    if header in SANITIZATION_FIELD_LABELS
                    else CELL_RUNTIME_POPULATED
                )
            _record_contract_cell(contract, sheet, coordinate, category, value, label)


def build_phase_rows(parser: SummaryParser) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for phase in parser.phases.keys():
        steps = parser.phase_steps(phase)
        timestamps = [str(step.get("ts")) for step in steps if step.get("ts")]
        rows.append([phase, parser.phase_status(phase), len(steps), min(timestamps) if timestamps else "", max(timestamps) if timestamps else ""])
    return rows


def build_step_rows(parser: SummaryParser) -> List[List[Any]]:
    rows: List[List[Any]] = []
    base_keys = {"id", "status", "ts", "http"}
    for phase, step in parser.all_steps():
        metadata = {key: value for key, value in step.items() if key not in base_keys}
        rows.append(
            [
                phase,
                step.get("id", NOT_FOUND),
                step.get("status", NOT_FOUND),
                step.get("ts", ""),
                step.get("http", ""),
                json.dumps(metadata, ensure_ascii=False, sort_keys=True) if metadata else "",
            ]
        )
    return rows


def build_artifact_rows(entries: List[FileEntry]) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for entry in entries:
        if entry.is_sensitive:
            continue
        rows.append([entry.phase, entry.relative_source_path, entry.http_status_file, entry.related_step, entry.description])
    return rows


def build_manifest_rows(loader: EvidenceRunLoader) -> List[List[Any]]:
    rows: List[List[Any]] = []
    manifest = manifest_data(loader.canonical_manifest_path())
    for key in sorted(manifest):
        rows.append([key, normalize_cell(manifest[key])])
    return rows


def checklist_rows(loader: EvidenceRunLoader, spec, summary_row: Dict[str, Any]) -> List[Dict[str, Any]]:
    parser = loader.parser or SummaryParser(loader.summary)
    kind = parser.detect_workflow_kind()

    def status_for(steps: List[Tuple[str, str]], require_all: bool = False) -> Tuple[str, str, str]:
        matched = []
        for phase, step_id in steps:
            step = parser.get_step(phase, step_id)
            if step:
                matched.append((phase, step))
        if not matched:
            return "not_found", "", ""
        ok_count = sum(1 for _, step in matched if step.get("status") == "ok")
        ok = ok_count == len(steps) if require_all else ok_count > 0
        phase, step = matched[-1]
        artifact = step.get("artifact") or step.get("manifest") or ""
        return ("ok" if ok else "fail", artifact, f"{phase}.{step.get('id')}")

    mappings = {
        "provider authentication": [("phase0", "jwt_provider"), ("phase1", "jwt_provider")],
        "consumer authentication": [("phase0", "jwt_consumer"), ("phase2", "jwt_consumer"), ("phase3", "jwt_consumer")],
        "asset created/published": [("phase1", "create_asset"), ("phase1b", "get_asset")],
        "policy created": [("phase1", "create_access_policy"), ("phase1", "create_contract_policy"), ("phase1b", "create_access_policy"), ("phase1b", "create_contract_policy")],
        "contract definition created": [("phase1", "create_contract_definition"), ("phase1b", "create_contract_definition")],
        "catalog discovery": [("phase1", "self_catalog"), ("phase2", "remote_catalog")],
        "offer policy selected/validated": [("phase2", "offer_policy_valid"), ("phase2", "catalog_asset_found")],
        "contract negotiation started": [("phase2", "contract_negotiation_started")],
        "negotiation finalized / agreement obtained": [("phase2", "negotiation_finalized"), ("phase2", "get_contract_agreement")],
        "transfer started/completed": [("phase3", "transfer_started"), ("phase3", "transfer_final_state"), ("phase3b", "transfer_started"), ("phase3b", "transfer_final_state")],
        "data downloaded/fetched": [("phase3", "data_consumed"), ("phase4", "save_download"), ("phase4b", "storage_fetch")],
        "manifest/hash generated": [("phase4", "save_download"), ("phase4b", "storage_fetch")],
    }
    rows = []
    for requirement in CHECKLIST_REQUIREMENTS:
        status, artifact, evidence_step = status_for(mappings[requirement])
        notes = ""
        if requirement == "transfer started/completed" and summary_row["overall_status"] == "PASS_WITH_NOTE":
            status = "ok"
            notes = summary_row["notes"]
        if requirement == "manifest/hash generated" and summary_row.get("sha256") not in ("", NOT_FOUND):
            status = "ok"
            artifact = summary_row.get("manifest_json", artifact)
        rows.append(
            {
                "test_id": spec.test_id,
                "requirement": requirement,
                "status": status,
                "evidence_file": artifact,
                "evidence_step": evidence_step,
                "notes": notes or (f"Workflow kind: {kind}" if status == "ok" else ""),
            }
        )
    return rows


def package_manifest_rows(entries: List[FileEntry], loader: EvidenceRunLoader, include_assets: bool = False) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for entry in entries:
        rows.append(
            {
                "test_id": entry.test_id,
                "file_path": entry.relative_source_path,
                "category": entry.category,
                "include_in_package": entry.include_in_package,
                "reason": entry.exclusion_reason or ("recommended" if entry.include_in_package else "excluded"),
                "size_bytes": entry.source_path.stat().st_size if entry.source_path.exists() and not entry.is_sensitive else 0,
            }
        )
    manifest = loader.canonical_manifest_path()
    if manifest and manifest.exists():
        rows.append(
            {
                "test_id": loader.spec.test_id,
                "file_path": relative_to_repo(loader.repo_root, manifest),
                "category": "manifest",
                "include_in_package": True,
                "reason": "canonical download manifest",
                "size_bytes": manifest.stat().st_size,
            }
        )
    if include_assets:
        asset = loader.latest_asset_path()
        if asset and asset.exists():
            rows.append(
                {
                    "test_id": loader.spec.test_id,
                    "file_path": relative_to_repo(loader.repo_root, asset),
                    "category": "asset",
                    "include_in_package": True,
                    "reason": "downloaded asset",
                    "size_bytes": asset.stat().st_size,
                }
            )
    return rows


def apply_workbook_format(wb: Workbook) -> None:
    status_fills = {
        "PASS": "D8EAD2",
        "ok": "D8EAD2",
        "PASS_WITH_NOTE": "FFF2CC",
        "PARTIAL": "FFF2CC",
        "FAIL": "F4CCCC",
        "fail": "F4CCCC",
        "INCOMPLETE": "F4CCCC",
    }
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
        for row in ws.iter_rows():
            for cell in row:
                if cell.row == 1 or (isinstance(cell.value, str) and cell.value in TABLE_TITLES):
                    cell.font = Font(bold=True)
                if isinstance(cell.value, str) and cell.value in status_fills:
                    cell.fill = PatternFill("solid", fgColor=status_fills[cell.value])
                if cell.column_letter in {"Y", "F"} or cell.value and isinstance(cell.value, str) and len(cell.value) > 80:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx in range(1, ws.max_column + 1):
            max_len = 8
            for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx):
                for item in cell:
                    max_len = max(max_len, min(len(str(item.value or "")), 60))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


def write_rows(
    ws,
    headers: List[str],
    rows: List[Dict[str, Any]],
    config: Dict[str, Any],
    repo_root: Path,
    sanitize: bool,
    redact_paths: bool,
    contract: Optional[WorkbookContract] = None,
) -> None:
    sheet = ws.title
    for index, header in enumerate(headers, 1):
        cell = ws.cell(1, index, header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAD3")
        _record_contract_cell(contract, sheet, cell.coordinate, CELL_PROTECTED_TEMPLATE_TEXT, header, header)
    for row_index, row in enumerate(rows, 2):
        values = [
            sanitize_value(row.get(header, ""), config, repo_root, sanitize, redact_paths)
            for header in headers
        ]
        for index, (header, value) in enumerate(zip(headers, values), 1):
            ws.cell(row_index, index, value)
            category = (
                CELL_SANITIZATION_ALLOWED
                if header in SANITIZATION_FIELD_LABELS
                else CELL_RUNTIME_POPULATED
            )
            _record_contract_cell(
                contract,
                sheet,
                f"{get_column_letter(index)}{row_index}",
                category,
                value,
                header,
            )


def resolve_output_path(
    repo_root: Path,
    output: Optional[str],
    export_dir: Optional[str],
    timestamp: Optional[str],
    timestamp_suffix: bool,
) -> Path:
    ts = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    if output:
        path = resolve_repo_path(repo_root, output)
    elif export_dir:
        return resolve_repo_path(repo_root, export_dir) / f"ippcp_evidence_summary_{ts}.xlsx"
    else:
        raise ValueError("--output or --export-dir is required")
    if not timestamp_suffix:
        return path
    return path.with_name(f"{path.stem}_{ts}{path.suffix}")


def _resolve_policy_flags(args, config: Dict[str, Any]) -> Tuple[bool, bool]:
    defaults = config.get("defaults") or {}
    profiles = config.get("profiles") or {}
    profile = profiles.get(args.profile or "default") or {}
    sanitize = args.sanitize_connectors
    if sanitize is None:
        sanitize = bool(profile.get("sanitize_connectors", defaults.get("sanitize_connectors", True)))
    redact = args.redact_local_paths
    if redact is None:
        redact = bool(profile.get("redact_local_paths", defaults.get("redact_local_paths", True)))
    return sanitize, redact


def _public_suffix(spec) -> str:
    return "<run-id>" if is_minimal_profile(spec) else spec.suffix


def main() -> int:
    args = parse_args()
    repo_root = Path(args.repo_root).resolve() if args.repo_root else find_repo_root(Path(__file__).resolve())
    config_path = resolve_repo_path(repo_root, args.config) if args.config else None
    config = load_test_config(config_path) if config_path else {"tests": {}, "presets": {}}
    overrides = parse_tests_override(args.tests)
    only_tests = parse_only_tests(args.only_tests)
    if args.tests and args.preset:
        print("WARN: --tests takes precedence; --preset is ignored", file=sys.stderr)
    try:
        specs = build_test_specs(
            config,
            overrides,
            only_tests,
            preset=None if overrides else args.preset,
        )
    except ValueError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    defaults = config.get("defaults") or {}
    evidence_dir = resolve_repo_path(repo_root, args.evidence_dir or defaults.get("evidence_dir") or "evidencias/runs")
    downloads_dir = resolve_repo_path(repo_root, args.downloads_dir or defaults.get("downloads_dir") or "downloads")
    warnings: List[str] = []
    sanitize_connectors, redact_local_paths = _resolve_policy_flags(args, config)

    if not specs:
        print(NO_SLOTS_SELECTED_ERROR, file=sys.stderr)
        return 1

    try:
        registry = load_asset_registry(config)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    jobs: List[Tuple[Any, EvidenceRunLoader, Any]] = []
    for spec in specs:
        loader = EvidenceRunLoader(repo_root, evidence_dir, downloads_dir, spec)
        if not loader.exists():
            message = f"{spec.test_id}: missing run summary at {loader.summary_path}"
            print(f"ERROR: {message}", file=sys.stderr)
            return 1
        try:
            loader.load(include_env=False)
            loader, spec, _asset = classify_loader(loader, registry)
        except ClassificationError as exc:
            print(str(exc), file=sys.stderr)
            return 1
        except Exception as exc:
            print(f"ERROR: {spec.test_id}: cannot load evidence: {exc}", file=sys.stderr)
            return 1
        if not is_minimal_profile(spec):
            loader.env = loader.load_env_files()
        validation = validate_classified_run(loader, spec)
        if args.strict and not validation.ok:
            print(
                f"ERROR: {spec.test_id}: semantic validation failed: {', '.join(validation.findings)}",
                file=sys.stderr,
            )
            return 1
        for message in ConnectorSanitizer.validate_workflow_roles(spec, config):
            warn(message, warnings, args.verbose)
        jobs.append((spec, loader, validation))

    minimal_specs = [spec for spec, _, _ in jobs if is_minimal_profile(spec)]
    delivery_specs = [spec for spec, _, _ in jobs if not is_minimal_profile(spec)]
    minimal_only = bool(minimal_specs) and not delivery_specs
    publication = package_publication_status([spec for spec, _, _ in jobs])
    print(f"package.publication_ready={str(publication['publication_ready']).lower()}")
    if publication["publication_blockers"]:
        print("package.publication_blockers:", file=sys.stderr)
        for blocker in publication["publication_blockers"]:
            print(f"  {blocker}", file=sys.stderr)

    wb = Workbook()
    wb.remove(wb.active)
    contract = WorkbookContract()
    summary_rows: List[Dict[str, Any]] = []
    slot_rows: List[Dict[str, Any]] = []
    raw_index_rows: List[Dict[str, Any]] = []
    checklist_all: List[Dict[str, Any]] = []
    package_rows: List[Dict[str, Any]] = []
    indexer = FileIndexer(repo_root)
    publication_sheets = set()

    for spec, loader, validation in jobs:
        parser = loader.parser or SummaryParser(loader.summary)
        if is_minimal_profile(spec):
            try:
                model = extract_minimal_publication_model(loader, spec)
            except Exception as exc:
                print(f"ERROR: {spec.test_id}: cannot build publication model: {exc}", file=sys.stderr)
                return 1
            if validation.status != "not_recorded":
                # Keep the public projection allowlisted; record validator in notes only.
                pass
            row = minimal_publication_summary_row(model, spec)
            summary_rows.append(row)
            publication_raw, publication_checklist, publication_package = minimal_publication_aggregate_rows(model, spec.sheet_name[:31])
            raw_index_rows.extend(publication_raw)
            checklist_all.extend(publication_checklist)
            package_rows.extend(publication_package)
            append_minimal_publication_sheet(wb, spec, model, contract=contract)
            publication_sheets.add(spec.sheet_name[:31])
            slot_rows.append(
                slot_inventory_row(spec, suffix_value=_public_suffix(spec), validation=validation)
            )
            continue

        entries = indexer.collect(loader.run_dir, spec, parser)
        row = build_summary_row(loader, spec, repo_root)
        if validation.findings:
            note = row.get("notes") or ""
            extra = f"semantic_validation={validation.status}"
            row["notes"] = f"{note}; {extra}".strip("; ")
        summary_rows.append(row)
        ws = wb.create_sheet(spec.sheet_name[:31])
        append_table(
            ws,
            "Datos generales",
            ["field", "value"],
            [
                ["slot", spec.test_id],
                ["asset", spec.display_name],
                ["asset_key", spec.asset_key],
                ["family", spec.family],
                ["variant", spec.variant],
                ["transport", spec.transport],
            ["critical", spec.critical],
            ["publication_profile", spec.publication_profile],
            ["publication_safe", spec.publication_safe],
            ["test_id", spec.test_id],
                ["workflow", spec.workflow],
                ["asset_type", spec.asset_type],
                ["asset_config", spec.asset_config],
                ["ds_name", loader.summary.get("ds_name", NOT_FOUND)],
                ["started_at", loader.summary.get("started_at", NOT_FOUND)],
                ["provider_connector", spec.provider_connector],
                ["consumer_connector", spec.consumer_connector],
            ],
            contract=contract,
        )
        append_table(
            ws,
            "IDs principales",
            ["field", "value"],
            [[key, row[key]] for key in SUMMARY_COLUMNS if key.endswith("_id") or key in {"asset_id", "sha256"}],
            contract=contract,
        )
        append_table(ws, "Estado por fase", ["phase", "status", "step_count", "first_ts", "last_ts"], build_phase_rows(parser), contract=contract)
        append_table(ws, "Pasos (summary)", ["phase", "step_id", "status", "ts", "http", "metadata"], build_step_rows(parser), contract=contract)
        append_table(ws, "Artefactos por fase", ["phase", "artifact", "http_status", "related_step", "description"], build_artifact_rows(entries), contract=contract)
        append_table(ws, "Manifest", ["field", "value"], build_manifest_rows(loader), contract=contract)
        append_table(
            ws,
            "Notas de interpretación",
            ["field", "value"],
            [["overall_status", row["overall_status"]], ["notes", row["notes"]]],
            contract=contract,
        )
        for entry in entries:
            raw_index_rows.append(
                {
                    "test_id": entry.test_id,
                    "suffix": entry.suffix,
                    "phase": entry.phase,
                    "file_path": entry.relative_source_path,
                    "file_name": entry.file_name,
                    "is_sensitive": entry.is_sensitive,
                    "include_in_package": entry.include_in_package,
                    "description": entry.description,
                    "http_status_file": entry.http_status_file,
                    "related_step": entry.related_step,
                }
            )
        checklist_all.extend(checklist_rows(loader, spec, row))
        package_rows.extend(package_manifest_rows(entries, loader))
        slot_rows.append(
            slot_inventory_row(spec, suffix_value=_public_suffix(spec), validation=validation)
        )

    ws_slots = wb.create_sheet("Slot Map", 0)
    write_rows(ws_slots, SLOT_MAP_COLUMNS, slot_rows, config, repo_root, sanitize_connectors, redact_local_paths, contract=contract)
    ws_summary = wb.create_sheet("Summary", 1)
    write_rows(ws_summary, SUMMARY_COLUMNS, summary_rows, config, repo_root, sanitize_connectors, redact_local_paths, contract=contract)
    ws_raw = wb.create_sheet("Raw JSON Index")
    write_rows(
        ws_raw,
        ["test_id", "suffix", "phase", "file_path", "file_name", "is_sensitive", "include_in_package", "description", "http_status_file", "related_step"],
        raw_index_rows,
        config,
        repo_root,
        sanitize_connectors,
        redact_local_paths,
        contract=contract,
    )
    ws_check = wb.create_sheet("Evidence Checklist")
    write_rows(ws_check, ["test_id", "requirement", "status", "evidence_file", "evidence_step", "notes"], checklist_all, config, repo_root, sanitize_connectors, redact_local_paths, contract=contract)
    ws_package = wb.create_sheet("Package Manifest")
    write_rows(ws_package, ["test_id", "file_path", "category", "include_in_package", "reason", "size_bytes"], package_rows, config, repo_root, sanitize_connectors, redact_local_paths, contract=contract)

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in (None, "") or not isinstance(cell.value, str):
                    continue
                new_value = sanitize_value(
                    cell.value, config, repo_root, sanitize_connectors, redact_local_paths
                )
                if new_value != cell.value:
                    cell.value = new_value

    apply_workbook_format(wb)
    final_cells = workbook_cell_snapshot(wb)
    expected_publication_cells: Dict[str, Dict[str, Any]] = {}
    if minimal_only:
        expected_publication_cells = final_cells
    else:
        for spec in minimal_specs:
            sheet_name = spec.sheet_name[:31]
            expected_publication_cells[sheet_name] = final_cells.get(sheet_name, {})

    audit_findings = audit_minimal_publication_workbook(
        wb,
        expected_cells=expected_publication_cells,
        protected_cells={},
        minimal_only=minimal_only,
        contract=contract,
        publication_sheets=publication_sheets if not minimal_only else set(final_cells),
    )
    output_buffer = io.BytesIO()
    if not audit_findings:
        wb.save(output_buffer)
        serialized = output_buffer.getvalue()
        publication_serialized_sheets = set(final_cells) if minimal_only else set(publication_sheets)
        audit_findings.extend(
            audit_minimal_publication_xlsx_bytes(
                serialized,
                publication_sheet_names=publication_serialized_sheets,
            )
        )
        if not audit_findings:
            loaded = load_workbook(io.BytesIO(serialized), data_only=False)
            audit_findings.extend(
                audit_minimal_publication_workbook(
                    loaded,
                    expected_cells=expected_publication_cells,
                    protected_cells={},
                    minimal_only=minimal_only,
                    contract=contract,
                    publication_sheets=publication_sheets if not minimal_only else set(final_cells),
                )
            )

    out = resolve_output_path(
        repo_root,
        args.output,
        args.export_dir,
        args.timestamp,
        args.timestamp_suffix,
    )
    if minimal_specs and any(spec.suffix in out.name for spec in minimal_specs):
        audit_findings.append("runtime suffix of a minimal_publication slot appears in output filename")
    if audit_findings:
        if out.exists():
            out.unlink()
        for finding in audit_findings:
            print(f"ERROR: workbook audit: {finding}", file=sys.stderr)
        return 1
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_bytes(output_buffer.getvalue())
    print(f"Excel written: {relative_to_repo(repo_root, out)}")
    if minimal_only:
        print(
            "minimal_publication workbook passed the full audit; manual review is required "
            "before publication"
        )
    for row in summary_rows:
        print(f"{row.get('test_id')}: {row.get('overall_status')} ({row.get('suffix')}) {row.get('asset_key')}")

    if args.strict and warnings:
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
