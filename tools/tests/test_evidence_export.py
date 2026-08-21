from __future__ import annotations

import json
import io
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName


REPO_ROOT = Path(__file__).resolve().parents[2]
TOOLS_DIR = REPO_ROOT / "tools"
sys.path.insert(0, str(TOOLS_DIR))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from evidence_fixtures import (  # noqa: E402
    ASSET_FIXTURES,
    create_asset_run,
    tests_override,
)
from evidence_assets import classify_loader, load_asset_registry  # noqa: E402
from evidence_common import (  # noqa: E402
    EvidenceRunLoader,
    PublicationScanner,
    audit_minimal_publication_workbook,
    audit_minimal_publication_xlsx_bytes,
    build_test_specs,
    build_minimal_publication_model,
    extract_minimal_publication_model,
    load_test_config,
    parse_only_tests,
    parse_tests_override,
    validate_allowlisted_json,
    workbook_cell_snapshot,
)
from export_evidence_to_excel import SUMMARY_COLUMNS, minimal_publication_execution_outcome  # noqa: E402
from package_evidence_bundle import (  # noqa: E402
    MANIFEST_FIELDS,
    MINIMAL_PUBLICATION_FILES,
    MINIMAL_PUBLICATION_JSON_SCHEMAS,
    PACKAGE_ROOT,
    build_minimal_publication_documents,
)


CONFIG_PATH = TOOLS_DIR / "evidence_export.tests.yaml"
PACKAGE_SCRIPT = TOOLS_DIR / "package_evidence_bundle.py"
EXCEL_SCRIPT = TOOLS_DIR / "export_evidence_to_excel.py"

CANARIES = (
    "CANARY-API-KEY",
    "CANARY-AUTHORIZATION",
    "CANARY-PASSWORD",
    "CANARY-PAYLOAD",
    "CANARY-PREVIEW",
    "CANARY-PHASE-ENV",
    "CANARY-DATA-ADDRESS",
    "canary.internal.invalid",
    "/Users/example/private",
    "9" * 10,
    "eyJFAKECANARYTOKEN0123456789",
)


class EvidenceExportTest(unittest.TestCase):
    def setUp(self) -> None:
        self.config = load_test_config(CONFIG_PATH)

    def specs(self, tests: str | None = None, only_tests: str | None = None, preset: str | None = None):
        return build_test_specs(
            self.config,
            parse_tests_override(tests),
            parse_only_tests(only_tests),
            preset=preset,
        )

    def run_cli(self, script: Path, *args: str) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            [sys.executable, str(script), *args],
            cwd=REPO_ROOT,
            text=True,
            capture_output=True,
            check=False,
        )

    def write_json(self, path: Path, data: object) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(json.dumps(data, indent=2) + "\n", encoding="utf-8")

    def create_minimal_publication_fixture(self, root: Path, suffix: str = "synthetic-t4") -> Path:
        run = create_asset_run(
            root,
            "ingestion_api_v2",
            suffix=suffix,
            with_canaries=True,
        )
        self.write_json(
            run / "phase4" / "40_data_response.json",
            {"business_payload": "CANARY-PAYLOAD"},
        )
        self.write_json(
            run / "phase4" / "41_data_preview.json",
            {"preview": "CANARY-PREVIEW"},
        )
        self.write_json(
            run / "phase4" / "download_manifest.json",
            {
                "bytes": 1234,
                "sha256": "a" * 64,
                "private_url": "https://canary.internal.invalid/data",
                "asset_slug": "ippcp_ingesta_api_pull_pre_api_key",
                "content_kind": "json",
                "extension": "json",
                "media_type": "application/json",
                "transfer_type": "HttpData-PULL",
            },
        )
        return run

    def create_delivery_fixtures(self, root: Path) -> None:
        create_asset_run(root, "csv_b2_legacy")
        create_asset_run(root, "wfs_ciudad")
        create_asset_run(root, "sparql")

    def delivery_tests_arg(self) -> str:
        return tests_override(
            {
                "T1": ASSET_FIXTURES["csv_b2_legacy"]["suffix"],
                "T2": ASSET_FIXTURES["wfs_ciudad"]["suffix"],
                "T3": ASSET_FIXTURES["sparql"]["suffix"],
            }
        )

    def create_minimal_publication_workbook(self, root: Path, output: Path) -> subprocess.CompletedProcess[str]:
        return self.run_cli(
            EXCEL_SCRIPT,
            "--repo-root",
            str(root),
            "--config",
            str(CONFIG_PATH),
            "--only-tests",
            "T4",
            "--tests",
            "T4=synthetic-t4",
            "--output",
            str(output),
        )

    def rewrite_xlsx(
        self, content: bytes, additions: dict[str, bytes]
    ) -> bytes:
        output = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(content)) as source:
            with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as target:
                for info in source.infolist():
                    if info.filename not in additions:
                        target.writestr(info, source.read(info.filename))
                for name, data in additions.items():
                    target.writestr(name, data)
        return output.getvalue()

    def assert_workbook_safety_contract(
        self, workbook_path: Path, expected_sheet_count: int
    ) -> None:
        workbook = load_workbook(workbook_path, data_only=False)
        self.assertEqual(len(workbook.sheetnames), expected_sheet_count)
        self.assertEqual(list(workbook.defined_names), [])
        for worksheet in workbook.worksheets:
            self.assertEqual(worksheet.sheet_state, "visible")
            self.assertFalse(
                any(dimension.hidden for dimension in worksheet.row_dimensions.values())
            )
            self.assertFalse(
                any(
                    dimension.hidden
                    for dimension in worksheet.column_dimensions.values()
                )
            )
            self.assertFalse(
                any(
                    cell.comment or cell.hyperlink
                    for row in worksheet.iter_rows()
                    for cell in row
                )
            )
        with zipfile.ZipFile(workbook_path) as archive:
            relationship_parts = {
                info.filename
                for info in archive.infolist()
                if info.filename.endswith(".rels")
            }
            self.assertEqual(
                relationship_parts,
                {"_rels/.rels", "xl/_rels/workbook.xml.rels"},
            )
            relationship_text = "\n".join(
                archive.read(name).decode("utf-8")
                for name in relationship_parts
            )
            self.assertNotIn('TargetMode="External"', relationship_text)
            self.assertNotIn("externalLink", relationship_text)

    def test_runtime_selection_modes(self) -> None:
        self.assertEqual([spec.test_id for spec in self.specs()], [])
        self.assertEqual(self.specs(only_tests="T4"), [])
        self.assertEqual(
            [spec.test_id for spec in self.specs("T4=synthetic-t4")],
            ["T4"],
        )
        self.assertEqual(
            [spec.test_id for spec in self.specs("T4=synthetic-t4", "T4")],
            ["T4"],
        )
        self.assertEqual(
            [spec.test_id for spec in self.specs("T1=a,T3=c")],
            ["T1", "T3"],
        )
        self.assertEqual(
            [spec.test_id for spec in self.specs(preset="legacy_assessment")],
            ["T1", "T2", "T3"],
        )
        self.assertEqual(
            [spec.test_id for spec in self.specs(preset="legacy_assessment", only_tests="T1")],
            ["T1"],
        )
        self.assertEqual(
            [spec.test_id for spec in self.specs("T1=runtime", preset="legacy_assessment")],
            ["T1"],
        )

    def test_package_cli_selection_matches_shared_contract(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            fixture_root = Path(tmp)
            self.create_delivery_fixtures(fixture_root)
            self.create_minimal_publication_fixture(fixture_root)
            cases = (
                (("--tests", self.delivery_tests_arg()), {"T1", "T2", "T3"}),
                (
                    ("--only-tests", "T4", "--tests", "T4=synthetic-t4"),
                    {"T4"},
                ),
                (
                    (
                        "--tests",
                        f"{self.delivery_tests_arg()},T4=synthetic-t4",
                    ),
                    {"T1", "T2", "T3", "T4"},
                ),
            )
            for arguments, expected in cases:
                with self.subTest(expected=expected):
                    result = self.run_cli(
                        PACKAGE_SCRIPT,
                        "--repo-root",
                        str(fixture_root),
                        "--config",
                        str(CONFIG_PATH),
                        "--dry-run",
                        *arguments,
                    )
                    self.assertEqual(result.returncode, 0, result.stderr)
                    selected = {
                        line.split("\t")[1]
                        for line in result.stdout.splitlines()
                        if line.count("\t") == 3
                        and line.split("\t")[1].startswith("T")
                    }
                    self.assertEqual(selected, expected)

    def test_slot_defaults_do_not_bind_assets(self) -> None:
        self.assertEqual(self.config["tests"], {})
        preset = self.config["presets"]["legacy_assessment"]["tests"]
        self.assertEqual(list(preset), ["T1", "T2", "T3"])
        self.assertTrue(all(isinstance(suffix, str) and suffix for suffix in preset.values()))
        assets = self.config["assets"]
        self.assertEqual(
            set(assets),
            {
                "ingestion_api_v2",
                "csv_b2_legacy",
                "wfs_juntas",
                "wfs_ciudad",
                "sparql",
            },
        )
        self.assertTrue(assets["ingestion_api_v2"]["critical"])
        self.assertEqual(
            assets["ingestion_api_v2"]["publication_profile"],
            "minimal_publication",
        )
        self.assertFalse(assets["csv_b2_legacy"]["critical"])
        self.assertEqual(assets["csv_b2_legacy"]["publication_profile"], "standard")
        self.assertTrue(assets["ingestion_api_v2"]["publication_safe"])
        self.assertFalse(assets["csv_b2_legacy"]["publication_safe"])
        self.assertEqual(
            SUMMARY_COLUMNS[:9],
            [
                "test_id",
                "asset_key",
                "display_name",
                "family",
                "variant",
                "transport",
                "critical",
                "publication_profile",
                "publication_safe",
            ],
        )
        self.assertEqual(
            MANIFEST_FIELDS,
            [
                "test_id",
                "suffix",
                "source_path",
                "zip_path",
                "category",
                "size_bytes",
                "sha256",
                "sanitized",
                "included",
                "exclusion_reason",
            ],
        )

    def test_default_t1_t3_package_inventory_shape(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            fixture_root = Path(tmp)
            self.create_delivery_fixtures(fixture_root)
            expected_folders = {
                "T1": "T1_ingesta_csv",
                "T2": "T2_wfs_ciudad",
                "T3": "T3_sparql",
            }
            result = self.run_cli(
                PACKAGE_SCRIPT,
                "--repo-root",
                str(fixture_root),
                "--config",
                str(CONFIG_PATH),
                "--tests",
                self.delivery_tests_arg(),
                "--dry-run",
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            rows = [line.split("\t") for line in result.stdout.splitlines() if line.count("\t") == 3]
            selected = {row[1] for row in rows if row[1].startswith("T")}
            included_targets = {row[3] for row in rows if row[0] == "INCLUDE" and row[1].startswith("T")}
            self.assertEqual(selected, {"T1", "T2", "T3"})
            for slot, folder in expected_folders.items():
                self.assertTrue(
                    any(f"{PACKAGE_ROOT}/{folder}/summary.json" == target for target in included_targets),
                    folder,
                )
            self.assertTrue(all("T4" not in row for row in rows))

    def test_only_t4_without_suffix_fails_before_evidence_read(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            for script, output_args in (
                (PACKAGE_SCRIPT, ("--dry-run",)),
                (EXCEL_SCRIPT, ("--output", str(Path(tmp) / "must-not-exist.xlsx"))),
            ):
                with self.subTest(script=script.name):
                    result = self.run_cli(
                        script,
                        "--repo-root",
                        tmp,
                        "--config",
                        str(CONFIG_PATH),
                        "--only-tests",
                        "T4",
                        *output_args,
                    )
                    self.assertEqual(result.returncode, 1)
                    self.assertIn("no slots selected", result.stderr)
                    self.assertIn("Supply --tests SLOT=SUFFIX", result.stderr)
                    self.assertNotIn("missing run summary", result.stderr)

    def test_minimal_publication_strict_package_uses_exact_allowlists_and_excludes_canaries(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            fixture_root = Path(tmp)
            self.create_minimal_publication_fixture(fixture_root)
            output = fixture_root / "t4-publication.zip"
            result = self.run_cli(
                PACKAGE_SCRIPT,
                "--repo-root",
                str(fixture_root),
                "--config",
                str(CONFIG_PATH),
                "--only-tests",
                "T4",
                "--tests",
                "T4=synthetic-t4",
                "--output",
                str(output),
                "--strict",
            )
            self.assertEqual(result.returncode, 0, result.stderr)

            expected_entries = {
                f"{PACKAGE_ROOT}/README_PACKAGE.txt",
                f"{PACKAGE_ROOT}/package_manifest.json",
                f"{PACKAGE_ROOT}/package_manifest.csv",
                f"{PACKAGE_ROOT}/package_status.json",
                f"{PACKAGE_ROOT}/slot_inventory.json",
                *{
                    f"{PACKAGE_ROOT}/T4_ingestion_api/{file_name}"
                    for file_name in MINIMAL_PUBLICATION_FILES
                },
            }
            with zipfile.ZipFile(output) as archive:
                actual_entries = {
                    info.filename for info in archive.infolist() if not info.is_dir()
                }
                self.assertEqual(actual_entries, expected_entries)
                archive_bytes = b"\n".join(
                    archive.read(name) for name in sorted(actual_entries)
                )
                archive_text = archive_bytes.decode("utf-8", errors="ignore")
                for canary in CANARIES:
                    self.assertNotIn(canary, archive_text)
                self.assertEqual(PublicationScanner.findings(archive_text), [])
                self.assertNotIn("phase1_env.sh", archive_text)
                self.assertNotIn("40_data_response.json", archive_text)
                self.assertNotIn("41_data_preview.json", archive_text)

                for file_name, schema in MINIMAL_PUBLICATION_JSON_SCHEMAS.items():
                    data = json.loads(
                        archive.read(
                            f"{PACKAGE_ROOT}/T4_ingestion_api/{file_name}"
                        )
                    )
                    self.assertEqual(validate_allowlisted_json(data, schema), [])

                manifest = json.loads(
                    archive.read(f"{PACKAGE_ROOT}/package_manifest.json")
                )
                status = json.loads(
                    archive.read(f"{PACKAGE_ROOT}/package_status.json")
                )
                self.assertTrue(status["publication_ready"])
                self.assertEqual(status["publication_blockers"], [])
                self.assertEqual(
                    status["slots"][0]["publication_profile"],
                    "minimal_publication",
                )
                self.assertTrue(status["slots"][0]["publication_safe"])
                self.assertNotIn("critical asset", json.dumps(status))
                for row in manifest:
                    self.assertEqual(set(row), set(MANIFEST_FIELDS))
                    self.assertFalse(row["source_path"])
                    self.assertIn(row["suffix"], {"", "<run-id>"})

    def test_unknown_publication_field_is_rejected(self) -> None:
        schema = MINIMAL_PUBLICATION_JSON_SCHEMAS["sanitized_manifest.json"]
        invalid = {
            "schema_version": "1.0",
            "artifact_type": "sanitized-download-manifest",
            "download": {
                "status": "ok",
                "byte_count": 1,
                "sha256_algorithm": "SHA-256",
                "sha256_verified": True,
                "sha256_value": "<withheld-pending-publication-approval>",
                "payload_included": False,
                "unknown": "must fail",
            },
            "delivery": {
                "mode": "download",
                "operation": "not_recorded",
                "http_method": "not_recorded",
                "http_status": "not_recorded",
                "manifest_kind": "not_recorded",
                "request_body_persisted": False,
                "response_body_persisted": False,
                "download_persisted": False,
            },
        }
        self.assertTrue(validate_allowlisted_json(invalid, schema))

    def test_publication_scanner_detects_generic_canaries(self) -> None:
        self.assertIn(
            "generic_canary",
            PublicationScanner.findings("prefix CANARY-UNLISTED-SURFACE suffix"),
        )
        self.assertIn(
            "generic_canary",
            PublicationScanner.findings("prefix CANARY_UNLISTED_SURFACE suffix"),
        )

    def test_technical_capability_is_independent_of_execution_outcome(self) -> None:
        model = build_minimal_publication_model(
            test_id="T4",
            asset_type="synthetic",
            evidence_role="additional_validation",
            technical_provider_connector="PROVIDER",
            technical_consumer_connector="CONSUMER",
            phase_statuses={
                "phase0": "ok",
                "phase1": "ok",
                "phase2": "failed",
                "phase3": "ok",
                "phase4": "ok",
            },
            download_status="ok",
            byte_count=1,
            sha256_verified=True,
        )
        self.assertEqual(model.technical_status, "Validated")
        self.assertEqual(minimal_publication_execution_outcome(model)[0], "FAIL")

    def test_minimal_publication_xlsx_export_has_public_safe_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            fixture_root = Path(tmp)
            self.create_minimal_publication_fixture(fixture_root)
            output = fixture_root / "t4-publication.xlsx"
            result = self.create_minimal_publication_workbook(fixture_root, output)
            self.assertEqual(result.returncode, 0, result.stderr)
            workbook = load_workbook(output, data_only=False)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "Slot Map",
                    "Summary",
                    "T4_ingestion_api",
                    "Raw JSON Index",
                    "Evidence Checklist",
                    "Package Manifest",
                ],
            )
            self.assertTrue(
                all(sheet.sheet_state == "visible" for sheet in workbook.worksheets)
            )
            summary = {
                cell.value: workbook["Summary"].cell(2, cell.column).value
                for cell in workbook["Summary"][1]
            }
            self.assertEqual(summary["workflow"], "Ingestion API v2")
            self.assertEqual(summary["suffix"], "<run-id>")
            self.assertEqual(summary["asset_id"], "<asset-id>")
            self.assertEqual(summary["vocab_id"], "not_applicable")
            self.assertEqual(summary["bytes"], 1234)
            self.assertEqual(
                summary["sha256"], "<withheld-pending-publication-approval>"
            )
            self.assertEqual(summary["overall_status"], "PASS")
            detail_values = {
                row[0].value: row[1].value
                for row in workbook["T4_ingestion_api"].iter_rows(min_col=1, max_col=2)
                if row[0].value
            }
            self.assertEqual(detail_values["technical_status"], "Validated")
            self.assertNotIn("publication_eligibility", detail_values)
            workbook_text = "\n".join(
                str(cell.value)
                for sheet in workbook.worksheets
                for row in sheet.iter_rows()
                for cell in row
                if cell.value is not None
            )
            for canary in CANARIES:
                self.assertNotIn(canary, workbook_text)
            self.assertNotIn("synthetic-t4", workbook_text)
            self.assertNotIn("phase1_env.sh", workbook_text)
            self.assertNotIn("40_data_response.json", workbook_text)
            self.assert_workbook_safety_contract(output, 6)

    def test_default_and_mixed_workbook_contracts(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            fixture_root = Path(tmp)
            self.create_delivery_fixtures(fixture_root)
            default_output = fixture_root / "default.xlsx"
            default_result = self.run_cli(
                EXCEL_SCRIPT,
                "--repo-root",
                str(fixture_root),
                "--config",
                str(CONFIG_PATH),
                "--tests",
                self.delivery_tests_arg(),
                "--output",
                str(default_output),
            )
            self.assertEqual(default_result.returncode, 0, default_result.stderr)
            self.assertIn("package.publication_ready=false", default_result.stdout)
            self.assertIn("slot T1 uses standard_internal", default_result.stderr)
            default = load_workbook(default_output, data_only=False)
            self.assertEqual(
                default.sheetnames,
                [
                    "Slot Map",
                    "Summary",
                    "T1_ingesta_csv",
                    "T2_wfs_ciudad",
                    "T3_sparql",
                    "Raw JSON Index",
                    "Evidence Checklist",
                    "Package Manifest",
                ],
            )
            self.assert_workbook_safety_contract(default_output, 8)

            self.create_minimal_publication_fixture(fixture_root)
            mixed_output = fixture_root / "mixed.xlsx"
            mixed_result = self.run_cli(
                EXCEL_SCRIPT,
                "--repo-root",
                str(fixture_root),
                "--config",
                str(CONFIG_PATH),
                "--tests",
                f"{self.delivery_tests_arg()},T4=synthetic-t4",
                "--output",
                str(mixed_output),
            )
            self.assertEqual(mixed_result.returncode, 0, mixed_result.stderr)
            self.assertIn("package.publication_ready=false", mixed_result.stdout)
            self.assertIn("slot T1 uses standard_internal", mixed_result.stderr)
            self.assertNotIn("contains a critical", mixed_result.stderr)
            self.assertNotIn("internal and not publication-ready as a whole", mixed_result.stderr)
            mixed = load_workbook(mixed_output, data_only=False)
            self.assertEqual(len(mixed.sheetnames), 9)
            self.assertEqual(mixed.sheetnames[:5], default.sheetnames[:5])
            self.assert_workbook_safety_contract(mixed_output, 9)
            for sheet_name in (
                "T1_ingesta_csv",
                "T2_wfs_ciudad",
                "T3_sparql",
            ):
                for coordinate, value in workbook_cell_snapshot(default)[sheet_name].items():
                    self.assertEqual(
                        mixed[sheet_name][coordinate].value,
                        value,
                        f"{sheet_name}!{coordinate}",
                    )

    def test_bundle_and_workbook_project_same_minimal_publication_model(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            fixture_root = Path(tmp)
            self.create_minimal_publication_fixture(fixture_root)
            spec = self.specs("T4=synthetic-t4", "T4")[0]
            loader = EvidenceRunLoader(
                fixture_root,
                fixture_root / "evidencias" / "runs",
                fixture_root / "downloads",
                spec,
            ).load(include_env=False)
            loader, spec, _asset = classify_loader(loader, load_asset_registry(self.config))
            model = extract_minimal_publication_model(loader, spec)
            self.assertFalse(hasattr(model, "documents"))
            self.assertEqual(model.technical_status, "Validated")
            projected_documents = build_minimal_publication_documents(model, spec)
            archive_path = fixture_root / "publication.zip"
            package_result = self.run_cli(
                PACKAGE_SCRIPT,
                "--repo-root",
                str(fixture_root),
                "--config",
                str(CONFIG_PATH),
                "--only-tests",
                "T4",
                "--tests",
                "T4=synthetic-t4",
                "--output",
                str(archive_path),
                "--strict",
            )
            self.assertEqual(package_result.returncode, 0, package_result.stderr)
            workbook_path = fixture_root / "publication.xlsx"
            workbook_result = self.create_minimal_publication_workbook(
                fixture_root, workbook_path
            )
            self.assertEqual(workbook_result.returncode, 0, workbook_result.stderr)

            with zipfile.ZipFile(archive_path) as archive:
                summary_document = json.loads(
                    archive.read(
                        f"{PACKAGE_ROOT}/T4_ingestion_api/sanitized_summary.json"
                    )
                )
                manifest_document = json.loads(
                    archive.read(
                        f"{PACKAGE_ROOT}/T4_ingestion_api/sanitized_manifest.json"
                    )
                )
                validation_document = json.loads(
                    archive.read(
                        f"{PACKAGE_ROOT}/T4_ingestion_api/validation_status.json"
                    )
                )
            self.assertEqual(
                projected_documents,
                {
                    "sanitized_summary.json": summary_document,
                    "sanitized_manifest.json": manifest_document,
                    "validation_status.json": validation_document,
                },
            )
            workbook = load_workbook(workbook_path, data_only=False)
            headers = [cell.value for cell in workbook["Summary"][1]]
            row = dict(zip(headers, [cell.value for cell in workbook["Summary"][2]]))
            detail = {
                cells[0].value: cells[1].value
                for cells in workbook["T4_ingestion_api"].iter_rows(
                    min_col=1, max_col=2
                )
                if cells[0].value
            }
            self.assertEqual(row["test_id"], summary_document["test_id"])
            self.assertEqual(row["workflow"], model.public_flow_label)
            self.assertEqual(row["asset_type"], summary_document["asset_type"])
            self.assertEqual(
                row["provider_connector"],
                summary_document["technical_topology"][
                    "technical_provider_connector"
                ],
            )
            self.assertEqual(
                row["consumer_connector"],
                summary_document["technical_topology"][
                    "technical_consumer_connector"
                ],
            )
            self.assertEqual(
                row["asset_id"],
                summary_document["execution_identifiers"]["asset_id"],
            )
            self.assertEqual(
                row["download_status"], manifest_document["download"]["status"]
            )
            self.assertEqual(
                row["bytes"], manifest_document["download"]["byte_count"]
            )
            self.assertEqual(
                row["sha256"], manifest_document["download"]["sha256_value"]
            )
            self.assertIn(
                validation_document["semantic_validation"]["status"],
                row["notes"],
            )
            self.assertEqual(row["overall_status"], "PASS")
            self.assertEqual(detail["technical_status"], model.technical_status)
            self.assertEqual(detail["evidence_role"], model.evidence_role)
            self.assertEqual(
                detail["technical_provider_connector"],
                model.technical_provider_connector,
            )
            self.assertEqual(
                detail["technical_consumer_connector"],
                model.technical_consumer_connector,
            )
            for identifier, value in model.execution_identifiers.items():
                self.assertEqual(detail[identifier], value)
            for phase, status in model.phase_statuses.items():
                self.assertEqual(detail[phase], status)
            self.assertEqual(detail["delivery_mode"], model.delivery_mode)
            self.assertEqual(detail["operation"], model.http_operation)
            self.assertEqual(detail["http_method"], model.http_method)
            self.assertEqual(detail["http_status"], model.http_status)
            self.assertEqual(detail["manifest_kind"], model.manifest_kind)
            self.assertEqual(detail["download_status"], model.download_status)
            self.assertEqual(detail["byte_count"], model.byte_count)
            self.assertEqual(detail["sha256_algorithm"], model.sha256_algorithm)
            self.assertEqual(detail["sha256_verified"], model.sha256_verified)
            self.assertEqual(detail["sha256_value"], model.sha256_value)
            self.assertEqual(detail["payload_included"], model.payload_included)
            self.assertEqual(
                detail["semantic_validation_status"],
                model.semantic_validation_status,
            )
            self.assertEqual(
                detail["semantic_validation_source"],
                model.semantic_validation_source,
            )
            expected_logical_paths = {
                "T4_ingestion_api/sanitized_summary.json",
                "T4_ingestion_api/sanitized_manifest.json",
                "T4_ingestion_api/validation_status.json",
            }
            self.assertEqual(
                {
                    cell.value
                    for cell in workbook["Raw JSON Index"]["D"][1:]
                },
                expected_logical_paths,
            )
            self.assertEqual(
                {
                    cell.value
                    for cell in workbook["Package Manifest"]["B"][1:]
                },
                expected_logical_paths,
            )

    def test_minimal_publication_runtime_suffix_is_rejected_in_output_name(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            fixture_root = Path(tmp)
            self.create_minimal_publication_fixture(fixture_root)
            output = fixture_root / "report-synthetic-t4.xlsx"
            result = self.create_minimal_publication_workbook(fixture_root, output)
            self.assertEqual(result.returncode, 1)
            self.assertIn("runtime suffix of a minimal_publication slot appears in output filename", result.stderr)
            self.assertFalse(output.exists())

    def test_minimal_publication_runtime_suffix_is_rejected_in_package_name(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            fixture_root = Path(tmp)
            self.create_minimal_publication_fixture(fixture_root)
            output = fixture_root / "report-synthetic-t4.zip"
            output.write_bytes(b"pre-existing rejected output")
            result = self.run_cli(
                PACKAGE_SCRIPT,
                "--repo-root",
                str(fixture_root),
                "--config",
                str(CONFIG_PATH),
                "--only-tests",
                "T4",
                "--tests",
                "T4=synthetic-t4",
                "--output",
                str(output),
            )
            self.assertEqual(result.returncode, 1)
            self.assertIn("runtime suffix of a minimal_publication slot appears in output filename", result.stderr)
            self.assertFalse(output.exists())

    def test_minimal_publication_export_dir_name_is_stable_and_suffix_free(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            fixture_root = Path(tmp)
            self.create_minimal_publication_fixture(fixture_root)
            export_dir = fixture_root / "exports"
            result = self.run_cli(
                EXCEL_SCRIPT,
                "--repo-root",
                str(fixture_root),
                "--config",
                str(CONFIG_PATH),
                "--tests",
                "T4=synthetic-t4",
                "--timestamp",
                "20260820_120000",
                "--export-dir",
                str(export_dir),
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            outputs = list(export_dir.glob("*.xlsx"))
            self.assertEqual(
                [output.name for output in outputs],
                ["ippcp_evidence_summary_20260820_120000.xlsx"],
            )
            self.assertNotIn("synthetic-t4", outputs[0].name)

    def test_in_memory_ooxml_audit_rejects_malicious_surfaces(self) -> None:
        def workbook() -> tuple[Workbook, dict[str, dict[str, object]]]:
            result = Workbook()
            result.active.title = "Publication"
            result["Publication"]["A1"] = "safe"
            return result, workbook_cell_snapshot(result)

        mutations = {
            "hidden_sheet": lambda wb: setattr(wb["Publication"], "sheet_state", "hidden"),
            "hidden_row": lambda wb: setattr(wb["Publication"].row_dimensions[1], "hidden", True),
            "hidden_column": lambda wb: setattr(wb["Publication"].column_dimensions["A"], "hidden", True),
            "formula": lambda wb: setattr(wb["Publication"]["A1"], "value", "=1+1"),
            "comment": lambda wb: setattr(wb["Publication"]["A1"], "comment", Comment("note", "author")),
            "hyperlink": lambda wb: setattr(wb["Publication"]["A1"], "hyperlink", "https://example.invalid"),
            "defined_name": lambda wb: wb.defined_names.add(
                DefinedName("unsafe_name", attr_text="Publication!$A$1")
            ),
            "property": lambda wb: setattr(
                wb.properties, "description", "OOXML-CANARY"
            ),
            "unexpected_cell": lambda wb: setattr(wb["Publication"]["B1"], "value", "extra"),
        }
        for name, mutate in mutations.items():
            with self.subTest(surface=name):
                candidate, expected = workbook()
                mutate(candidate)
                findings = audit_minimal_publication_workbook(
                    candidate,
                    expected_cells=expected,
                    minimal_only=True,
                    canaries={"OOXML-CANARY"},
                )
                self.assertTrue(findings, name)

    def test_serialized_ooxml_audit_rejects_malicious_parts(self) -> None:
        workbook = Workbook()
        workbook.active.title = "Publication"
        workbook["Publication"]["A1"] = "safe"
        buffer = io.BytesIO()
        workbook.save(buffer)
        clean = buffer.getvalue()
        self.assertEqual(
            audit_minimal_publication_xlsx_bytes(clean, publication_sheet_names={"Publication"}),
            [],
        )
        external_relationship = b"""<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" Target="https://example.invalid" TargetMode="External"/>
</Relationships>"""
        attacks = {
            "external_relationship": {
                "xl/worksheets/_rels/sheet1.xml.rels": external_relationship
            },
            "macro": {"xl/vbaProject.bin": b"unsafe"},
            "custom_xml": {"customXml/item1.xml": b"<unsafe/>"},
            "unexpected_part": {"xl/unreviewed.xml": b"<unsafe/>"},
            "shared_string_canary": {
                "xl/sharedStrings.xml": (
                    b'<?xml version="1.0" encoding="UTF-8"?>'
                    b'<sst xmlns="http://schemas.openxmlformats.org/'
                    b'spreadsheetml/2006/main"><si><t>OOXML-CANARY</t></si></sst>'
                )
            },
        }
        for name, additions in attacks.items():
            with self.subTest(surface=name):
                malicious = self.rewrite_xlsx(clean, additions)
                findings = audit_minimal_publication_xlsx_bytes(
                    malicious,
                    publication_sheet_names={"Publication"},
                    canaries={"OOXML-CANARY"},
                )
                self.assertTrue(findings, name)


if __name__ == "__main__":
    unittest.main()
