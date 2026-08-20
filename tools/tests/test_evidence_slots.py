from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
TOOLS_DIR = REPO_ROOT / "tools"
sys.path.insert(0, str(TOOLS_DIR))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from evidence_fixtures import (  # noqa: E402
    ASSET_FIXTURES,
    complete_override,
    create_ambiguous_run,
    create_asset_run,
    create_unknown_run,
    tests_override,
)
from evidence_common import PublicationScanner, load_test_config  # noqa: E402


CONFIG_PATH = TOOLS_DIR / "evidence_export.tests.yaml"
EXCEL_SCRIPT = TOOLS_DIR / "export_evidence_to_excel.py"
PACKAGE_SCRIPT = TOOLS_DIR / "package_evidence_bundle.py"

SECRET_CANARIES = (
    "CANARY-API-KEY",
    "CANARY-AUTHORIZATION",
    "CANARY-PASSWORD",
    "CANARY-DATA-ADDRESS",
    "eyJFAKECANARYTOKEN0123456789",
    "Bearer CANARY-AUTHORIZATION",
)


class EvidenceSlotAssetTest(unittest.TestCase):
    def run_cli(self, script: Path, *args: str) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            [sys.executable, str(script), *args],
            cwd=REPO_ROOT,
            text=True,
            capture_output=True,
            check=False,
        )

    def export_excel(self, root: Path, tests: str, only: str | None = None, strict: bool = True):
        output = root / "out.xlsx"
        args = [
            "--repo-root",
            str(root),
            "--config",
            str(CONFIG_PATH),
            "--tests",
            tests,
            "--output",
            str(output),
        ]
        if only:
            args.extend(["--only-tests", only])
        if strict:
            args.append("--strict")
        result = self.run_cli(EXCEL_SCRIPT, *args)
        return result, output

    def export_zip(self, root: Path, tests: str, only: str | None = None, strict: bool = True):
        output = root / "out.zip"
        args = [
            "--repo-root",
            str(root),
            "--config",
            str(CONFIG_PATH),
            "--tests",
            tests,
            "--output",
            str(output),
        ]
        if only:
            args.extend(["--only-tests", only])
        if strict:
            args.append("--strict")
        result = self.run_cli(PACKAGE_SCRIPT, *args)
        return result, output

    def slot_map(self, workbook_path: Path) -> list[dict[str, object]]:
        workbook = load_workbook(workbook_path, data_only=False)
        headers = [cell.value for cell in workbook["Slot Map"][1]]
        return [
            dict(zip(headers, [cell.value for cell in row]))
            for row in workbook["Slot Map"].iter_rows(min_row=2, max_row=workbook["Slot Map"].max_row)
            if row[0].value
        ]

    def prepare_complete(self, root: Path, order: list[str]) -> str:
        for key in order:
            canaries = key == "ingestion_api_v2"
            create_asset_run(root, key, with_canaries=canaries)
        return complete_override(order)

    def test_complete_canonical_order(self) -> None:
        order = ["ingestion_api_v2", "wfs_juntas", "wfs_ciudad", "sparql"]
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            tests = self.prepare_complete(root, order)
            result, output = self.export_excel(root, tests)
            self.assertEqual(result.returncode, 0, result.stderr)
            rows = self.slot_map(output)
            self.assertEqual([row["slot"] for row in rows], ["T1", "T2", "T3", "T4"])
            self.assertEqual(
                [row["asset_key"] for row in rows],
                order,
            )
            self.assertEqual(
                load_workbook(output).sheetnames[2:6],
                ["T1_ingestion_api", "T2_wfs_juntas", "T3_wfs_ciudad", "T4_sparql"],
            )
            self.assertTrue(rows[0]["critical"])
            self.assertEqual(rows[0]["publication_profile"], "minimal_publication")
            self.assertFalse(rows[1]["critical"])
            zip_result, zip_output = self.export_zip(root, tests)
            self.assertEqual(zip_result.returncode, 0, zip_result.stderr)
            with zipfile.ZipFile(zip_output) as archive:
                names = {info.filename for info in archive.infolist() if not info.is_dir()}
                self.assertIn("ippcp_evidence_package/slot_inventory.json", names)
                inventory = json.loads(archive.read("ippcp_evidence_package/slot_inventory.json"))
                self.assertEqual([row["asset_key"] for row in inventory], order)
                status = json.loads(archive.read("ippcp_evidence_package/package_status.json"))
                self.assertFalse(status["publication_ready"])
                self.assertIn("slot T2 uses standard_internal", status["publication_blockers"])
                self.assertFalse(
                    any("critical asset" in blocker for blocker in status["publication_blockers"])
                )
                self.assertTrue(inventory[0]["publication_safe"])
                self.assertFalse(inventory[1]["publication_safe"])

    def test_complete_different_order(self) -> None:
        order = ["sparql", "wfs_ciudad", "ingestion_api_v2", "wfs_juntas"]
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            tests = self.prepare_complete(root, order)
            result, output = self.export_excel(root, tests)
            self.assertEqual(result.returncode, 0, result.stderr)
            rows = self.slot_map(output)
            self.assertEqual([row["asset_key"] for row in rows], order)
            self.assertEqual(rows[2]["slot"], "T3")
            self.assertTrue(rows[2]["critical"])
            self.assertEqual(
                load_workbook(output).sheetnames[2:6],
                ["T1_sparql", "T2_wfs_ciudad", "T3_ingestion_api", "T4_wfs_juntas"],
            )

    def test_complete_with_legacy_csv(self) -> None:
        order = ["csv_b2_legacy", "sparql", "wfs_juntas", "wfs_ciudad"]
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            tests = self.prepare_complete(root, order)
            result, output = self.export_excel(root, tests)
            self.assertEqual(result.returncode, 0, result.stderr)
            rows = self.slot_map(output)
            self.assertEqual(rows[0]["asset_key"], "csv_b2_legacy")
            self.assertEqual(rows[0]["publication_profile"], "standard")
            self.assertFalse(rows[0]["critical"])
            self.assertEqual(rows[0]["transport"], "InesDataStore")
            self.assertIn("T1_ingesta_csv", load_workbook(output).sheetnames)
            self.assertFalse(rows[0]["publication_safe"])

    def test_single_ingestion_same_policy_in_every_slot(self) -> None:
        policy_keys = (
            "asset_key",
            "family",
            "variant",
            "transport",
            "critical",
            "publication_profile",
            "publication_safe",
        )
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            policies = []
            for slot in ("T1", "T2", "T3", "T4"):
                suffix = f"synthetic-ingestion-{slot.lower()}"
                create_asset_run(
                    root, "ingestion_api_v2", suffix=suffix, with_canaries=True
                )
                result, output = self.export_excel(root, f"{slot}={suffix}")
                self.assertEqual(result.returncode, 0, result.stderr)
                rows = self.slot_map(output)
                self.assertEqual([row["slot"] for row in rows], [slot])
                row = rows[0]
                self.assertEqual(row["asset_key"], "ingestion_api_v2")
                self.assertTrue(row["critical"])
                self.assertEqual(row["publication_profile"], "minimal_publication")
                self.assertTrue(row["publication_safe"])
                policies.append({key: row[key] for key in policy_keys})
            self.assertEqual(policies[0], policies[1])
            self.assertEqual(policies[0], policies[2])
            self.assertEqual(policies[0], policies[3])

    def test_single_wfs_ciudad_in_t1(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_asset_run(root, "wfs_ciudad")
            result, output = self.export_excel(
                root, f"T1={ASSET_FIXTURES['wfs_ciudad']['suffix']}"
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            row = self.slot_map(output)[0]
            self.assertEqual(row["slot"], "T1")
            self.assertEqual(row["asset_key"], "wfs_ciudad")
            self.assertEqual(row["publication_profile"], "standard")
            self.assertFalse(row["critical"])
            self.assertFalse(row["publication_safe"])

    def test_wfs_and_sparql_detection_is_slot_independent(self) -> None:
        cases = (
            ("wfs_ciudad", ("T1", "T4")),
            ("wfs_juntas", ("T2", "T4")),
            ("sparql", ("T1", "T3")),
        )
        policy_keys = (
            "asset_key",
            "family",
            "variant",
            "transport",
            "critical",
            "publication_profile",
            "publication_safe",
        )
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            for asset_key, slots in cases:
                policies = []
                for slot in slots:
                    suffix = f"synthetic-{asset_key}-{slot.lower()}"
                    create_asset_run(root, asset_key, suffix=suffix)
                    result, output = self.export_excel(root, f"{slot}={suffix}")
                    self.assertEqual(result.returncode, 0, result.stderr)
                    row = self.slot_map(output)[0]
                    self.assertEqual(row["slot"], slot)
                    self.assertEqual(row["asset_key"], asset_key)
                    self.assertFalse(row["critical"])
                    self.assertEqual(row["publication_profile"], "standard")
                    self.assertFalse(row["publication_safe"])
                    policies.append({key: row[key] for key in policy_keys})
                self.assertEqual(policies[0], policies[1], asset_key)

    def test_partial_complete_does_not_autofill_slots(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_asset_run(root, "wfs_ciudad", suffix="synthetic-partial-t1")
            create_asset_run(root, "sparql", suffix="synthetic-partial-t3")
            result, output = self.export_excel(
                root, "T3=synthetic-partial-t3,T1=synthetic-partial-t1"
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            rows = self.slot_map(output)
            self.assertEqual([row["slot"] for row in rows], ["T1", "T3"])
            self.assertEqual(
                [row["asset_key"] for row in rows],
                ["wfs_ciudad", "sparql"],
            )
            sheets = load_workbook(output).sheetnames
            self.assertIn("T1_wfs_ciudad", sheets)
            self.assertIn("T3_sparql", sheets)
            self.assertNotIn("T2_wfs_juntas", sheets)
            self.assertFalse(any(name.startswith("T2_") for name in sheets))
            self.assertFalse(any(name.startswith("T4_") for name in sheets))

    def test_one_and_three_slot_complete_sets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_asset_run(root, "wfs_juntas", suffix="synthetic-one")
            one, one_out = self.export_excel(root, "T2=synthetic-one")
            self.assertEqual(one.returncode, 0, one.stderr)
            self.assertEqual([row["slot"] for row in self.slot_map(one_out)], ["T2"])
            create_asset_run(root, "wfs_ciudad", suffix="synthetic-three-t1")
            create_asset_run(root, "sparql", suffix="synthetic-three-t3")
            create_asset_run(root, "csv_b2_legacy", suffix="synthetic-three-t4")
            three, three_out = self.export_excel(
                root,
                "T1=synthetic-three-t1,T3=synthetic-three-t3,T4=synthetic-three-t4",
            )
            self.assertEqual(three.returncode, 0, three.stderr)
            self.assertEqual(
                [row["slot"] for row in self.slot_map(three_out)],
                ["T1", "T3", "T4"],
            )

    def test_four_ingestion_slots_are_publication_ready(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            parts = []
            for slot in ("T1", "T2", "T3", "T4"):
                suffix = f"synthetic-all-ingestion-{slot.lower()}"
                create_asset_run(root, "ingestion_api_v2", suffix=suffix, with_canaries=True)
                parts.append(f"{slot}={suffix}")
            zip_result, zip_path = self.export_zip(root, ",".join(parts))
            self.assertEqual(zip_result.returncode, 0, zip_result.stderr)
            self.assertIn("package.publication_ready=true", zip_result.stdout)
            with zipfile.ZipFile(zip_path) as archive:
                status = json.loads(archive.read("ippcp_evidence_package/package_status.json"))
            self.assertTrue(status["publication_ready"])
            self.assertEqual(status["publication_blockers"], [])
            self.assertTrue(all(slot["publication_safe"] for slot in status["slots"]))

    def test_only_tests_still_filters_selected_set(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_asset_run(root, "wfs_ciudad", suffix="synthetic-filter-t1")
            create_asset_run(root, "sparql", suffix="synthetic-filter-t3")
            result, output = self.export_excel(
                root,
                "T1=synthetic-filter-t1,T3=synthetic-filter-t3",
                only="T3",
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            self.assertEqual([row["slot"] for row in self.slot_map(output)], ["T3"])
            self.assertEqual(self.slot_map(output)[0]["asset_key"], "sparql")

    def test_single_wfs_juntas_in_t3(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_asset_run(root, "wfs_juntas")
            result, output = self.export_excel(
                root, f"T3={ASSET_FIXTURES['wfs_juntas']['suffix']}"
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            row = self.slot_map(output)[0]
            self.assertEqual(row["slot"], "T3")
            self.assertEqual(row["asset_key"], "wfs_juntas")

    def test_single_sparql_in_t2(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_asset_run(root, "sparql")
            result, output = self.export_excel(
                root, f"T2={ASSET_FIXTURES['sparql']['suffix']}"
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            row = self.slot_map(output)[0]
            self.assertEqual(row["slot"], "T2")
            self.assertEqual(row["asset_key"], "sparql")
            self.assertEqual(row["transport"], "HttpData-PULL")

    def test_missing_suffix_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            result, output = self.export_excel(Path(tmp), "T2=does-not-exist")
            self.assertEqual(result.returncode, 1)
            self.assertIn("missing run summary", result.stderr)
            self.assertFalse(output.exists())

    def test_incomplete_run_fails_in_strict(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_asset_run(root, "sparql", suffix="synthetic-incomplete", incomplete=True)
            result, output = self.export_excel(root, "T1=synthetic-incomplete")
            self.assertEqual(result.returncode, 1)
            self.assertIn("semantic validation failed", result.stderr)
            self.assertFalse(output.exists())

    def test_unclassifiable_run_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_unknown_run(root)
            result, _ = self.export_excel(root, "T2=synthetic-unknown")
            self.assertEqual(result.returncode, 1)
            self.assertIn("Unable to classify evidence run.", result.stderr)
            self.assertIn("slot=T2", result.stderr)
            self.assertIn("suffix=synthetic-unknown", result.stderr)
            self.assertIn("detected=unknown", result.stderr)

    def test_ambiguous_run_fails(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_ambiguous_run(root)
            result, _ = self.export_excel(root, "T2=synthetic-ambiguous")
            self.assertEqual(result.returncode, 1)
            self.assertIn("Unable to classify evidence run.", result.stderr)
            self.assertIn("detected=ambiguous", result.stderr)

    def test_critical_asset_hides_secrets_in_workbook_and_bundle(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_asset_run(root, "ingestion_api_v2", with_canaries=True)
            tests = f"T1={ASSET_FIXTURES['ingestion_api_v2']['suffix']}"
            excel_result, excel_path = self.export_excel(root, tests)
            zip_result, zip_path = self.export_zip(root, tests)
            self.assertEqual(excel_result.returncode, 0, excel_result.stderr)
            self.assertEqual(zip_result.returncode, 0, zip_result.stderr)
            workbook = load_workbook(excel_path, data_only=False)
            text = "\n".join(
                str(cell.value)
                for sheet in workbook.worksheets
                for row in sheet.iter_rows()
                for cell in row
                if cell.value is not None
            )
            with zipfile.ZipFile(zip_path) as archive:
                bundle = "\n".join(
                    archive.read(name).decode("utf-8", errors="ignore")
                    for name in archive.namelist()
                    if not name.endswith("/")
                )
            for canary in SECRET_CANARIES:
                self.assertNotIn(canary, text)
                self.assertNotIn(canary, bundle)
            self.assertEqual(PublicationScanner.findings(text), [])
            self.assertEqual(PublicationScanner.findings(bundle), [])
            self.assertNotIn("/Users/example/private", text)
            self.assertNotIn("/Users/example/private", bundle)

    def test_non_critical_assets_keep_permitted_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_asset_run(root, "sparql")
            tests = f"T2={ASSET_FIXTURES['sparql']['suffix']}"
            result, output = self.export_zip(root, tests)
            self.assertEqual(result.returncode, 0, result.stderr)
            with zipfile.ZipFile(output) as archive:
                names = {info.filename for info in archive.infolist() if not info.is_dir()}
            self.assertTrue(any(name.endswith("/summary.json") for name in names))
            self.assertTrue(any("/phase4/" in name for name in names))

    def test_mixed_complete_and_order_and_detection(self) -> None:
        order = ["wfs_ciudad", "ingestion_api_v2", "sparql", "wfs_juntas"]
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            tests = self.prepare_complete(root, order)
            excel_result, excel_path = self.export_excel(root, tests)
            zip_result, zip_path = self.export_zip(root, tests)
            self.assertEqual(excel_result.returncode, 0, excel_result.stderr)
            self.assertEqual(zip_result.returncode, 0, zip_result.stderr)
            rows = self.slot_map(excel_path)
            self.assertEqual([row["slot"] for row in rows], ["T1", "T2", "T3", "T4"])
            self.assertEqual([row["asset_key"] for row in rows], order)
            self.assertFalse(rows[0]["critical"])
            self.assertTrue(rows[1]["critical"])
            workbook = load_workbook(excel_path, data_only=False)
            self.assertTrue(any("Datos generales" in str(cell.value) for row in workbook["T2_ingestion_api"].iter_rows() for cell in row))
            text = "\n".join(
                str(cell.value)
                for sheet in workbook.worksheets
                for row in sheet.iter_rows()
                for cell in row
                if cell.value is not None
            )
            for canary in SECRET_CANARIES:
                self.assertNotIn(canary, text)
            with zipfile.ZipFile(zip_path) as archive:
                names = archive.namelist()
            self.assertTrue(any("T2_ingestion_api/sanitized_summary.json" in name for name in names))
            self.assertTrue(any("T1_wfs_ciudad/summary.json" in name for name in names))

    def test_runtime_cells_are_contracted_not_ignored(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_asset_run(root, "csv_b2_legacy")
            create_asset_run(root, "sparql")
            result, output = self.export_excel(
                root,
                tests_override(
                    {
                        "T1": ASSET_FIXTURES["csv_b2_legacy"]["suffix"],
                        "T2": ASSET_FIXTURES["sparql"]["suffix"],
                    }
                ),
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            workbook = load_workbook(output, data_only=False)
            values = {
                row[0].value: row[1].value
                for row in workbook["T1_ingesta_csv"].iter_rows(min_col=1, max_col=2)
                if row[0].value
            }
            self.assertEqual(values["slot"], "T1")
            self.assertEqual(values["asset_key"], "csv_b2_legacy")
            self.assertIn("provider_connector", values)
            self.assertIn("consumer_connector", values)
            self.assertEqual(values["ds_name"], "synthetic")

    def test_no_real_suffixes_or_endpoints_in_fixtures(self) -> None:
        config = load_test_config(CONFIG_PATH)
        real_suffixes = set()
        for preset in (config.get("presets") or {}).values():
            tests = (preset or {}).get("tests") or {}
            for value in tests.values():
                suffix = value if isinstance(value, str) else (value or {}).get("suffix")
                if suffix:
                    real_suffixes.add(str(suffix))
        fixture_suffixes = {item["suffix"] for item in ASSET_FIXTURES.values()}
        self.assertTrue(fixture_suffixes.isdisjoint(real_suffixes))
        self.assertNotIn("https://urbanismo.geoslab.com", json.dumps(ASSET_FIXTURES))
        self.assertNotIn("idezar-sig.zaragoza.es", json.dumps(ASSET_FIXTURES))


if __name__ == "__main__":
    unittest.main()
