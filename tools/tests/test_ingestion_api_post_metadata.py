#!/usr/bin/env python3
"""POST metadata-only support for ingestion_api_v2 evidence tooling."""

from __future__ import annotations

import json
import subprocess
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path

TOOLS_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = TOOLS_DIR.parent
sys.path.insert(0, str(TOOLS_DIR))
sys.path.insert(0, str(TOOLS_DIR / "tests"))

from evidence_assets import (  # noqa: E402
    classify_loader,
    load_asset_registry,
    validate_classified_run,
)
from evidence_common import EvidenceRunLoader, TestSpec, load_test_config  # noqa: E402
from evidence_fixtures import (  # noqa: E402
    ASSET_FIXTURES,
    create_asset_run,
    create_ingestion_post_metadata_run,
)
from package_evidence_bundle import PublicationScanner  # noqa: E402

PYTHON = "/Users/jpardo/anaconda3/envs/data_spaces_31018/bin/python"
CONFIG_PATH = TOOLS_DIR / "evidence_export.tests.yaml"
PACKAGE_SCRIPT = TOOLS_DIR / "package_evidence_bundle.py"
EXPORT_SCRIPT = TOOLS_DIR / "export_evidence_to_excel.py"


class IngestionApiPostMetadataTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.config = load_test_config(CONFIG_PATH)
        cls.registry = load_asset_registry(cls.config)

    def _load_classified(self, root: Path, suffix: str, slot: str = "T1"):
        spec = TestSpec(test_id=slot, suffix=suffix, sheet_name=slot)
        loader = EvidenceRunLoader(
            root, root / "evidencias" / "runs", root / "downloads", spec
        ).load(include_env=False)
        return classify_loader(loader, self.registry)

    def test_pre_get_still_requires_payload_and_download_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_asset_run(root, "ingestion_api_v2")
            loader, spec, asset = self._load_classified(
                root, ASSET_FIXTURES["ingestion_api_v2"]["suffix"]
            )
            self.assertEqual(asset.key, "ingestion_api_v2")
            validation = validate_classified_run(loader, spec)
            self.assertTrue(validation.ok, validation.findings)
            # Removing download surfaces must fail GET path.
            (loader.run_dir / "phase4" / "download_manifest.json").unlink()
            (loader.run_dir / "phase4" / "downloaded.json").unlink()
            validation = validate_classified_run(loader, spec)
            self.assertFalse(validation.ok)
            self.assertIn("json_payload:missing", validation.findings)
            self.assertIn("manifest:missing", validation.findings)

    def test_prod_post_classifies_and_validates_without_download(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_ingestion_post_metadata_run(root)
            loader, spec, asset = self._load_classified(root, "synthetic-ingestion-post")
            self.assertEqual(asset.key, "ingestion_api_v2")
            self.assertEqual(spec.publication_profile, "minimal_publication")
            validation = validate_classified_run(loader, spec)
            self.assertTrue(validation.ok, validation.findings)
            self.assertEqual(validation.source, "ingestion-api-post-metadata-validator")
            self.assertFalse((loader.run_dir / "phase4" / "download_manifest.json").exists())
            self.assertFalse((loader.run_dir / "phase4" / "downloaded.json").exists())

    def test_circe_prod_slug_classifies(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_ingestion_post_metadata_run(
                root,
                suffix="synthetic-circe-post",
                slug="ippcp_ingesta_pull_circe_prod",
                asset_config="asset_configs/real/ingesta/ingesta_api_pull_circe_prod.json",
            )
            # Override asset_id inside manifest/summary to Circe prod id.
            summary_path = root / "evidencias/runs/synthetic-circe-post/summary.json"
            summary = json.loads(summary_path.read_text(encoding="utf-8"))
            for step in summary["phases"]["phase1"]["steps"]:
                if step.get("id") == "create_asset":
                    step["asset_id"] = "ippcp-ingesta-pull-circe-prod"
                    step["asset_slug"] = "ippcp_ingesta_pull_circe_prod"
            summary_path.write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")
            manifest_path = root / "evidencias/runs/synthetic-circe-post/phase4/post_manifest.json"
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            manifest["asset_id"] = "ippcp-ingesta-pull-circe-prod"
            manifest_path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")
            _loader, _spec, asset = self._load_classified(root, "synthetic-circe-post")
            self.assertEqual(asset.key, "ingestion_api_v2")

    def test_post_invalid_variants_fail(self) -> None:
        cases = [
            {"http_status": 500},
            {"request_body_persisted": True},
            {"response_body_persisted": True},
            {"download_persisted": True},
            {"include_post_result": False},
            {"include_post_manifest": False},
            {"manifest_kind": "download"},
        ]
        for index, kwargs in enumerate(cases):
            with self.subTest(case=kwargs):
                with tempfile.TemporaryDirectory() as tmp:
                    root = Path(tmp)
                    create_ingestion_post_metadata_run(
                        root, suffix=f"synthetic-bad-{index}", **kwargs
                    )
                    loader, spec, _asset = self._load_classified(
                        root, f"synthetic-bad-{index}"
                    )
                    # Wrong manifest_kind falls through to GET validator.
                    validation = validate_classified_run(loader, spec)
                    self.assertFalse(validation.ok, validation.findings)

    def test_post_package_excludes_bodies_and_auth_canaries(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_ingestion_post_metadata_run(root, with_canaries=True)
            output = root / "post-package.zip"
            result = subprocess.run(
                [
                    PYTHON,
                    str(PACKAGE_SCRIPT),
                    "--repo-root",
                    str(root),
                    "--config",
                    str(CONFIG_PATH),
                    "--tests",
                    "T1=synthetic-ingestion-post",
                    "--output",
                    str(output),
                    "--strict",
                ],
                cwd=str(REPO_ROOT),
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            with zipfile.ZipFile(output) as archive:
                names = {info.filename for info in archive.infolist() if not info.is_dir()}
                joined = "\n".join(
                    archive.read(name).decode("utf-8", errors="ignore") for name in sorted(names)
                )
                self.assertTrue(any(name.endswith("sanitized_manifest.json") for name in names))
                self.assertNotIn("request_body.secret.json", joined)
                self.assertNotIn("response_body.secret.json", joined)
                self.assertNotIn("CANARY-API-KEY", joined)
                self.assertNotIn("CANARY-AUTHORIZATION", joined)
                self.assertNotIn("CANARY-REQUEST-BODY", joined)
                self.assertNotIn("CANARY-RESPONSE-BODY", joined)
                self.assertNotIn("synthetic-ingestion-post", joined)
                self.assertEqual(PublicationScanner.findings(joined), [])
                manifest = json.loads(
                    next(
                        archive.read(name)
                        for name in names
                        if name.endswith("sanitized_manifest.json")
                    )
                )
                self.assertEqual(manifest["delivery"]["mode"], "post_metadata_only")
                self.assertEqual(manifest["delivery"]["operation"], "POST")
                self.assertEqual(manifest["download"]["status"], "not_applicable")
                self.assertFalse(manifest["download"]["sha256_verified"])
                self.assertFalse(manifest["delivery"]["download_persisted"])

    def test_post_excel_pass_without_download_claim(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            create_ingestion_post_metadata_run(root)
            output = root / "post.xlsx"
            result = subprocess.run(
                [
                    PYTHON,
                    str(EXPORT_SCRIPT),
                    "--repo-root",
                    str(root),
                    "--config",
                    str(CONFIG_PATH),
                    "--tests",
                    "T1=synthetic-ingestion-post",
                    "--output",
                    str(output),
                    "--strict",
                ],
                cwd=str(REPO_ROOT),
                capture_output=True,
                text=True,
                check=False,
            )
            self.assertEqual(result.returncode, 0, result.stderr)
            from openpyxl import load_workbook

            wb = load_workbook(output, data_only=True)
            headers = [cell.value for cell in next(wb["Summary"].iter_rows(min_row=1, max_row=1))]
            rows = list(wb["Summary"].iter_rows(min_row=2, values_only=True))
            self.assertTrue(rows)
            row = dict(zip(headers, rows[0]))
            self.assertEqual(row["overall_status"], "PASS")
            self.assertEqual(row["download_status"], "not_applicable")
            self.assertIn("delivery_mode=post_metadata_only", row["notes"])
            detail = {
                cells[0].value: cells[1].value
                for cells in wb["T1_ingestion_api"].iter_rows(min_col=1, max_col=2)
                if cells[0].value
            }
            self.assertEqual(detail["delivery_mode"], "post_metadata_only")
            self.assertEqual(detail["operation"], "POST")
            self.assertEqual(detail["http_status"], "200")
            self.assertEqual(detail["download_status"], "not_applicable")
            self.assertFalse(detail["sha256_verified"])


if __name__ == "__main__":
    unittest.main()
