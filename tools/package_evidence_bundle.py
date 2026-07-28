#!/usr/bin/env python3
"""Package IPPCP evidence runs into a safe, shareable ZIP bundle."""

from __future__ import annotations

import argparse
import csv
import json
import shutil
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

from evidence_common import (
    ConnectorSanitizer,
    EvidenceRunLoader,
    FileEntry,
    FileIndexer,
    MINIMAL_PUBLICATION_PROFILE,
    NOT_FOUND,
    PublicationScanner,
    SecretScanner,
    SummaryParser,
    T4PublicationModel,
    build_test_specs,
    compute_sha256,
    extract_t4_publication_model,
    find_repo_root,
    load_test_config,
    parse_only_tests,
    parse_tests_override,
    relative_to_repo,
    resolve_repo_path,
    validate_allowlisted_json,
)


PACKAGE_ROOT = "ippcp_evidence_package"
MINIMAL_PUBLICATION_FILES = {
    "sanitized_summary.json",
    "sanitized_manifest.json",
    "validation_status.json",
}
MINIMAL_PUBLICATION_JSON_SCHEMAS: Dict[str, Dict[str, Any]] = {
    "sanitized_summary.json": {
        "schema_version": None,
        "test_id": None,
        "flow_type": None,
        "asset_type": None,
        "evidence_role": None,
        "technical_topology": {
            "technical_provider_connector": None,
            "technical_consumer_connector": None,
        },
        "execution_identifiers": {
            "run_id": None,
            "asset_id": None,
            "contract_definition_id": None,
            "negotiation_id": None,
            "agreement_id": None,
            "transfer_process_id": None,
        },
        "phases": {
            "phase0": {"status": None},
            "phase1": {"status": None},
            "phase2": {"status": None},
            "phase3": {"status": None},
            "phase4": {"status": None},
        },
    },
    "sanitized_manifest.json": {
        "schema_version": None,
        "artifact_type": None,
        "download": {
            "status": None,
            "byte_count": None,
            "sha256_algorithm": None,
            "sha256_verified": None,
            "sha256_value": None,
            "payload_included": None,
        },
    },
    "validation_status.json": {
        "schema_version": None,
        "semantic_validation": {
            "status": None,
            "source": None,
        },
        "publication_checks": {
            "payload_excluded": None,
            "phase_environment_excluded": None,
            "raw_requests_responses_excluded": None,
            "identifiers_replaced": None,
            "hash_value_withheld": None,
        },
    },
}
MANIFEST_FIELDS = [
    "test_id",
    "suffix",
    "source_path",
    "zip_path",
    "category",
    "size_bytes",
    "sha256",
    "sanitized",
    "included",
    "exclusion_reason",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", required=True, help="YAML/JSON config with tests")
    parser.add_argument("--tests", help="Override tests as T1=SUFFIX,T2=SUFFIX")
    parser.add_argument("--only-tests", help="Select only these configured test IDs, comma-separated")
    parser.add_argument("--excel", help="Optional Excel report to include")
    parser.add_argument("--output", help="Output .zip path")
    parser.add_argument("--export-dir", help="Base directory for timestamped exports")
    parser.add_argument("--timestamp", help="Timestamp to use for generated names, format YYYYMMDD_HHMMSS")
    parser.add_argument("--timestamp-suffix", action="store_true", help="Append timestamp before .zip")
    parser.add_argument("--repo-root", help="Repository root")
    parser.add_argument("--sanitize-connectors", action="store_true", help="Apply connector aliases to packaged copies")
    parser.add_argument("--redact-local-paths", action="store_true", default=True, help="Redact absolute local repo paths")
    parser.add_argument("--no-redact-local-paths", dest="redact_local_paths", action="store_false", help="Do not redact local paths")
    parser.add_argument("--include-downloaded-assets", action="store_true", help="Include latest downloaded assets")
    parser.add_argument("--strict", action="store_true", help="Fail on warnings/audit findings")
    parser.add_argument("--dry-run", action="store_true", help="List planned package entries without writing ZIP")
    parser.add_argument("--verbose", action="store_true")
    return parser.parse_args()


def log(message: str, verbose: bool) -> None:
    if verbose:
        print(message)


def add_manifest_row(
    rows: List[Dict[str, Any]],
    *,
    test_id: str,
    suffix: str = "",
    source_path: str = "",
    zip_path: str = "",
    category: str,
    size_bytes: int = 0,
    sha256: str = "",
    sanitized: bool = False,
    included: bool,
    exclusion_reason: str = "",
) -> None:
    rows.append(
        {
            "test_id": test_id,
            "suffix": suffix,
            "source_path": source_path,
            "zip_path": zip_path,
            "category": category,
            "size_bytes": size_bytes,
            "sha256": sha256,
            "sanitized": sanitized,
            "included": included,
            "exclusion_reason": exclusion_reason,
        }
    )


def text_suffix(path: Path) -> bool:
    return path.suffix.lower() in {".json", ".http", ".txt", ".csv", ".md"}


def sanitize_secret_key_names(text: str) -> str:
    # Packaged copies should not expose credential key names even when values are redacted.
    replacements = {
        "secretAccessKey": "redactedSecretKey",
        "accessKeyId": "redactedAccessKeyId",
        "access_token": "redacted_access_token",
        "refresh_token": "redacted_refresh_token",
        "client_secret": "redacted_client_secret",
    }
    result = text
    for source, target in replacements.items():
        result = result.replace(source, target)
    return result


def copy_for_package(
    source: Path,
    dest: Path,
    repo_root: Path,
    config: Dict[str, Any],
    sanitize_connectors: bool,
    redact_local_paths: bool,
) -> bool:
    dest.parent.mkdir(parents=True, exist_ok=True)
    if text_suffix(source):
        text = source.read_text(encoding="utf-8", errors="replace")
        text = ConnectorSanitizer.apply(
            text,
            config.get("connector_aliases") if sanitize_connectors else {},
            redact_local_paths,
            repo_root,
        )
        text = sanitize_secret_key_names(text)
        dest.write_text(text, encoding="utf-8")
        return sanitize_connectors or redact_local_paths
    shutil.copy2(source, dest)
    return False


def package_path_for_entry(entry: FileEntry, test_folder: str) -> str:
    if entry.file_name == "summary.json":
        return f"{PACKAGE_ROOT}/{test_folder}/summary.json"
    rel = Path(entry.source_path.name)
    try:
        phase_rel = Path(*Path(entry.relative_source_path).parts[-2:])
        if phase_rel.parts[0].startswith("phase"):
            rel = phase_rel
    except Exception:
        pass
    return f"{PACKAGE_ROOT}/{test_folder}/{rel.as_posix()}"


def should_scan(entry: FileEntry) -> bool:
    return entry.source_path.suffix.lower() in {".json", ".http", ".txt", ".csv"}


def stage_entry(
    rows: List[Dict[str, Any]],
    entry: FileEntry,
    test_folder: str,
    staging_root: Path,
    repo_root: Path,
    config: Dict[str, Any],
    sanitize_connectors: bool,
    redact_local_paths: bool,
    strict: bool,
    warnings: List[str],
) -> None:
    zip_path = package_path_for_entry(entry, test_folder)
    if not entry.include_in_package:
        add_manifest_row(
            rows,
            test_id=entry.test_id,
            suffix=entry.suffix,
            source_path=entry.relative_source_path,
            category=entry.category,
            included=False,
            exclusion_reason=entry.exclusion_reason,
        )
        return

    if should_scan(entry):
        ok, reason = SecretScanner.scan_file(entry.source_path)
        if not ok:
            message = f"{entry.test_id}: excluding {entry.relative_source_path}: {reason}"
            warnings.append(message)
            add_manifest_row(
                rows,
                test_id=entry.test_id,
                suffix=entry.suffix,
                source_path=entry.relative_source_path,
                category=entry.category,
                included=False,
                exclusion_reason=reason,
            )
            return

    dest = staging_root / Path(zip_path).relative_to(PACKAGE_ROOT)
    sanitized = copy_for_package(entry.source_path, dest, repo_root, config, sanitize_connectors, redact_local_paths)
    add_manifest_row(
        rows,
        test_id=entry.test_id,
        suffix=entry.suffix,
        source_path=entry.relative_source_path,
        zip_path=zip_path,
        category=entry.category,
        size_bytes=dest.stat().st_size,
        sha256=compute_sha256(dest),
        sanitized=sanitized,
        included=True,
    )


def stage_external_file(
    rows: List[Dict[str, Any]],
    *,
    source: Path,
    zip_path: str,
    test_id: str,
    suffix: str,
    category: str,
    staging_root: Path,
    repo_root: Path,
    config: Dict[str, Any],
    sanitize_connectors: bool,
    redact_local_paths: bool,
    reason_if_missing: str = "missing",
) -> None:
    source_rel = relative_to_repo(repo_root, source)
    if not source.exists():
        add_manifest_row(rows, test_id=test_id, suffix=suffix, source_path=source_rel, category=category, included=False, exclusion_reason=reason_if_missing)
        return
    if text_suffix(source):
        ok, reason = SecretScanner.scan_file(source)
        if not ok:
            add_manifest_row(rows, test_id=test_id, suffix=suffix, source_path=source_rel, category=category, included=False, exclusion_reason=reason)
            return
    dest = staging_root / Path(zip_path).relative_to(PACKAGE_ROOT)
    sanitized = copy_for_package(source, dest, repo_root, config, sanitize_connectors, redact_local_paths)
    add_manifest_row(
        rows,
        test_id=test_id,
        suffix=suffix,
        source_path=source_rel,
        zip_path=zip_path,
        category=category,
        size_bytes=dest.stat().st_size,
        sha256=compute_sha256(dest),
        sanitized=sanitized,
        included=True,
    )


def build_minimal_publication_documents(
    model: T4PublicationModel,
) -> Dict[str, Dict[str, Any]]:
    """Project the neutral T4 model into the package JSON contract."""
    return {
        "sanitized_summary.json": {
            "schema_version": "1.0",
            "test_id": model.test_id,
            "flow_type": model.flow_type,
            "asset_type": model.asset_type,
            "evidence_role": model.evidence_role,
            "technical_topology": {
                "technical_provider_connector": model.technical_provider_connector,
                "technical_consumer_connector": model.technical_consumer_connector,
            },
            "execution_identifiers": dict(model.execution_identifiers),
            "phases": {
                phase: {"status": model.phase_statuses[phase]}
                for phase in ("phase0", "phase1", "phase2", "phase3", "phase4")
            },
        },
        "sanitized_manifest.json": {
            "schema_version": "1.0",
            "artifact_type": "sanitized-download-manifest",
            "download": {
                "status": model.download_status,
                "byte_count": model.byte_count,
                "sha256_algorithm": model.sha256_algorithm,
                "sha256_verified": model.sha256_verified,
                "sha256_value": model.sha256_value,
                "payload_included": model.payload_included,
            },
        },
        "validation_status.json": {
            "schema_version": "1.0",
            "semantic_validation": {
                "status": model.semantic_validation_status,
                "source": model.semantic_validation_source,
            },
            "publication_checks": {
                "payload_excluded": not model.payload_included,
                "phase_environment_excluded": True,
                "raw_requests_responses_excluded": True,
                "identifiers_replaced": True,
                "hash_value_withheld": True,
            },
        },
    }


def stage_minimal_publication(
    rows: List[Dict[str, Any]],
    loader: EvidenceRunLoader,
    spec,
    staging_root: Path,
) -> None:
    model = extract_t4_publication_model(loader, spec)
    documents = build_minimal_publication_documents(model)
    if set(documents) != MINIMAL_PUBLICATION_FILES:
        raise RuntimeError("minimal publication documents differ from the archive allowlist")

    for file_name in sorted(documents):
        data = documents[file_name]
        schema_findings = validate_allowlisted_json(data, MINIMAL_PUBLICATION_JSON_SCHEMAS[file_name])
        if schema_findings:
            raise RuntimeError(f"{spec.test_id} {file_name} violates field allowlist: {schema_findings}")
        text = json.dumps(data, ensure_ascii=False, indent=2) + "\n"
        publication_findings = PublicationScanner.findings(text)
        if publication_findings:
            raise RuntimeError(f"{spec.test_id} {file_name} violates publication policy: {publication_findings}")
        dest = staging_root / spec.sheet_name / file_name
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_text(text, encoding="utf-8")
        add_manifest_row(
            rows,
            test_id=spec.test_id,
            suffix="<run-id>",
            source_path="",
            zip_path=f"{PACKAGE_ROOT}/{spec.sheet_name}/{file_name}",
            category="publication_metadata",
            size_bytes=dest.stat().st_size,
            sha256=compute_sha256(dest),
            sanitized=True,
            included=True,
        )


def stage_excel(
    rows: List[Dict[str, Any]],
    excel_path: Path,
    staging_root: Path,
    repo_root: Path,
    config: Dict[str, Any],
    sanitize_connectors: bool,
    redact_local_paths: bool,
) -> None:
    zip_path = f"{PACKAGE_ROOT}/{excel_path.name}"
    dest = staging_root / excel_path.name
    dest.parent.mkdir(parents=True, exist_ok=True)
    sanitized = False
    if sanitize_connectors or redact_local_paths:
        wb = load_workbook(excel_path)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        new_value = ConnectorSanitizer.apply(
                            cell.value,
                            config.get("connector_aliases") if sanitize_connectors else {},
                            redact_local_paths,
                            repo_root,
                        )
                        if new_value != cell.value:
                            sanitized = True
                            cell.value = new_value
        wb.save(dest)
    else:
        shutil.copy2(excel_path, dest)
    add_manifest_row(
        rows,
        test_id="_package",
        source_path=relative_to_repo(repo_root, excel_path),
        zip_path=zip_path,
        category="excel",
        size_bytes=dest.stat().st_size,
        sha256=compute_sha256(dest),
        sanitized=sanitized,
        included=True,
    )


def write_package_readme(staging_root: Path, config: Dict[str, Any], specs, sanitize_connectors: bool) -> Path:
    minimal_only = bool(specs) and all(spec.publication_profile == MINIMAL_PUBLICATION_PROFILE for spec in specs)
    if minimal_only:
        contents = [
            "- Per-test sanitized summary, manifest metadata, and validation status.",
            "- package_manifest.json and package_manifest.csv list the allowlisted files.",
            "- No raw phase artifact, phase environment, source path, or downloaded payload is included.",
        ]
    else:
        contents = [
            "- Excel summary report, when provided with --excel.",
            "- Per-test evidence folders with summary.json, non-sensitive JSON artifacts, .http files, and download manifests.",
            "- package_manifest.json and package_manifest.csv list included and excluded files.",
        ]
    lines = [
        "IPPCP Evidence Package",
        f"Generated at: {datetime.now().isoformat(timespec='seconds')}",
        "",
        "Contents:",
        *contents,
        "",
        "Tests included:",
    ]
    for spec in specs:
        if spec.publication_profile == MINIMAL_PUBLICATION_PROFILE:
            provider = spec.technical_provider_connector
            consumer = spec.technical_consumer_connector
            suffix = "<run-id>"
        else:
            provider = spec.provider_connector
            consumer = spec.consumer_connector
            suffix = spec.suffix
        if sanitize_connectors:
            provider = ConnectorSanitizer.apply(provider, config.get("connector_aliases"), False)
            consumer = ConnectorSanitizer.apply(consumer, config.get("connector_aliases"), False)
        lines.append(f"- {spec.test_id}: workflow={spec.workflow}, suffix={suffix}, provider={provider}, consumer={consumer}")
    lines.extend(
        [
            "",
            "Security exclusions:",
            "- *.sensitive.json, *.secret.json, phase*_env.sh, runtime/env/**, flujos/*/user_*.sh, and *.body are not included.",
            "- Files with unredacted credential/JWT-like values are excluded.",
            "- Original evidencias/runs/ and downloads/ files are not modified by this packaging tool.",
            "",
            "Interpreting STARTED transfers:",
            "- In B1/B2 flows, a transfer may remain STARTED while the evidence still contains a verified download.",
            "- If the manifest or summary records bytes and sha256 for save_download/storage_fetch, the material data retrieval was verified.",
        ]
    )
    if sanitize_connectors:
        lines.append("")
        lines.append("Connector aliases applied:")
        lines.append("- Private connector names were replaced with public aliases.")
        for target in (config.get("connector_aliases") or {}).values():
            lines.append(f"- {target}")
    path = staging_root / "README_PACKAGE.txt"
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


def write_package_manifests(staging_root: Path, rows: List[Dict[str, Any]]) -> None:
    json_path = staging_root / "package_manifest.json"
    csv_path = staging_root / "package_manifest.csv"
    json_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=MANIFEST_FIELDS)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in MANIFEST_FIELDS})


def create_zip(staging_root: Path, output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for path in sorted(staging_root.rglob("*")):
            if path.is_file():
                archive.write(path, f"{PACKAGE_ROOT}/{path.relative_to(staging_root).as_posix()}")


def audit_zip(zip_path: Path, sanitize_connectors: bool) -> List[str]:
    findings: List[str] = []
    forbidden = [b"secretAccessKey", b"accessKeyId"]
    if sanitize_connectors:
        forbidden.extend([b"conn-erick-test3", b"conn-edgar-test3"])
    with zipfile.ZipFile(zip_path) as archive:
        for info in archive.infolist():
            name = info.filename
            if "sensitive" in name.lower() or name.endswith(".secret.json"):
                findings.append(f"forbidden file name in zip: {name}")
            data = archive.read(info)
            for token in forbidden:
                if token in data:
                    findings.append(f"forbidden string {token.decode()} in {name}")
            try:
                text = data.decode("utf-8", errors="ignore")
            except Exception:
                text = ""
            if SecretScanner.JWT_RE.search(text):
                findings.append(f"JWT-like token in {name}")
    return findings


def minimal_publication_archive_allowlist(spec) -> set[str]:
    entries = {
        f"{PACKAGE_ROOT}/{spec.sheet_name}/{file_name}"
        for file_name in MINIMAL_PUBLICATION_FILES
    }
    entries.update(
        {
            f"{PACKAGE_ROOT}/README_PACKAGE.txt",
            f"{PACKAGE_ROOT}/package_manifest.json",
            f"{PACKAGE_ROOT}/package_manifest.csv",
        }
    )
    return entries


def audit_minimal_publication_zip(zip_path: Path, spec) -> List[str]:
    findings: List[str] = []
    allowed_entries = minimal_publication_archive_allowlist(spec)
    with zipfile.ZipFile(zip_path) as archive:
        actual_entries = {info.filename for info in archive.infolist() if not info.is_dir()}
        for name in sorted(actual_entries - allowed_entries):
            findings.append(f"unknown archive entry: {name}")
        for name in sorted(allowed_entries - actual_entries):
            findings.append(f"missing archive entry: {name}")

        for name in sorted(actual_entries):
            data = archive.read(name)
            text = data.decode("utf-8", errors="ignore")
            for finding in PublicationScanner.findings(text):
                findings.append(f"{finding} in {name}")

            if name.endswith(".json"):
                try:
                    parsed = json.loads(text)
                except json.JSONDecodeError:
                    findings.append(f"invalid JSON in {name}")
                    continue
                file_name = Path(name).name
                if file_name in MINIMAL_PUBLICATION_JSON_SCHEMAS:
                    for finding in validate_allowlisted_json(
                        parsed, MINIMAL_PUBLICATION_JSON_SCHEMAS[file_name]
                    ):
                        findings.append(f"{finding} in {name}")
                elif file_name == "package_manifest.json":
                    if not isinstance(parsed, list):
                        findings.append(f"package manifest is not a list in {name}")
                    else:
                        for index, row in enumerate(parsed):
                            if not isinstance(row, dict):
                                findings.append(f"package manifest row {index} is not an object")
                                continue
                            unknown = set(row) - set(MANIFEST_FIELDS)
                            missing = set(MANIFEST_FIELDS) - set(row)
                            if unknown:
                                findings.append(f"package manifest row {index} unknown fields: {sorted(unknown)}")
                            if missing:
                                findings.append(f"package manifest row {index} missing fields: {sorted(missing)}")
                            if row.get("source_path"):
                                findings.append(f"source path present in package manifest row {index}")
                            if row.get("suffix") not in {"", "<run-id>"}:
                                findings.append(f"real suffix present in package manifest row {index}")
    return findings


def print_dry_run(rows: List[Dict[str, Any]]) -> None:
    for row in rows:
        marker = "INCLUDE" if row.get("included") else "EXCLUDE"
        target = row.get("zip_path") or row.get("exclusion_reason")
        print(f"{marker}\t{row.get('test_id')}\t{row.get('source_path')}\t{target}")


def resolve_output_path(
    repo_root: Path,
    output: Optional[str],
    export_dir: Optional[str],
    timestamp: Optional[str],
    timestamp_suffix: bool,
) -> Optional[Path]:
    ts = timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
    if output:
        path = resolve_repo_path(repo_root, output)
    elif export_dir:
        return resolve_repo_path(repo_root, export_dir) / f"ippcp_evidence_package_{ts}.zip"
    else:
        return None
    if not timestamp_suffix:
        return path
    return path.with_name(f"{path.stem}_{ts}{path.suffix}")


def main() -> int:
    args = parse_args()
    repo_root = Path(args.repo_root).resolve() if args.repo_root else find_repo_root(Path(__file__).resolve())
    config = load_test_config(resolve_repo_path(repo_root, args.config))
    only_tests = parse_only_tests(args.only_tests)
    specs = build_test_specs(config, parse_tests_override(args.tests), only_tests)
    defaults = config.get("defaults") or {}
    evidence_dir = resolve_repo_path(repo_root, defaults.get("evidence_dir") or "evidencias/runs")
    downloads_dir = resolve_repo_path(repo_root, defaults.get("downloads_dir") or "downloads")
    warnings: List[str] = []

    if not specs:
        if only_tests:
            print(
                "ERROR: selected tests have no configured or runtime suffix; "
                "supply --tests TEST_ID=SUFFIX",
                file=sys.stderr,
            )
        else:
            print("ERROR: no tests configured", file=sys.stderr)
        return 1
    minimal_specs = [spec for spec in specs if spec.publication_profile == MINIMAL_PUBLICATION_PROFILE]
    minimal_only = len(minimal_specs) == len(specs)
    if minimal_specs and not minimal_only:
        message = (
            "minimal_publication T4 is mixed with delivery profiles; "
            "use --only-tests T4 for a publication-ready T4 package"
        )
        warnings.append(message)
        print(f"WARN: {message}", file=sys.stderr)
    if minimal_specs and args.include_downloaded_assets:
        message = "downloaded assets are forbidden for minimal_publication and will not be staged for T4"
        warnings.append(message)
        print(f"WARN: {message}", file=sys.stderr)
    if minimal_specs and args.excel:
        message = "Excel inclusion is disabled for minimal_publication"
        warnings.append(message)
        print(f"WARN: {message}", file=sys.stderr)
    for spec in specs:
        for message in ConnectorSanitizer.validate_workflow_roles(spec, config):
            warnings.append(message)
            print(f"WARN: {message}", file=sys.stderr)

    staging_parent = Path(tempfile.mkdtemp(prefix="ippcp_bundle_"))
    staging_root = staging_parent / PACKAGE_ROOT
    staging_root.mkdir(parents=True, exist_ok=True)
    rows: List[Dict[str, Any]] = []
    indexer = FileIndexer(repo_root)
    try:
        for spec in specs:
            loader = EvidenceRunLoader(repo_root, evidence_dir, downloads_dir, spec)
            if not loader.exists():
                message = f"{spec.test_id}: missing run summary at {loader.summary_path}"
                warnings.append(message)
                print(f"WARN: {message}", file=sys.stderr)
                add_manifest_row(
                    rows,
                    test_id=spec.test_id,
                    suffix="<run-id>" if spec.publication_profile == MINIMAL_PUBLICATION_PROFILE else spec.suffix,
                    source_path="" if spec.publication_profile == MINIMAL_PUBLICATION_PROFILE else relative_to_repo(repo_root, loader.summary_path),
                    category="evidence",
                    included=False,
                    exclusion_reason="missing_run",
                )
                continue
            if spec.publication_profile == MINIMAL_PUBLICATION_PROFILE:
                loader.load(include_env=False)
                stage_minimal_publication(rows, loader, spec, staging_root)
                continue

            loader.load()
            parser = loader.parser or SummaryParser(loader.summary)
            entries = indexer.collect(loader.run_dir, spec, parser)
            for entry in entries:
                stage_entry(rows, entry, spec.sheet_name, staging_root, repo_root, config, args.sanitize_connectors, args.redact_local_paths, args.strict, warnings)
            manifest = loader.canonical_manifest_path()
            if manifest:
                stage_external_file(
                    rows,
                    source=manifest,
                    zip_path=f"{PACKAGE_ROOT}/{spec.sheet_name}/downloads/latest.manifest.json",
                    test_id=spec.test_id,
                    suffix=spec.suffix,
                    category="manifest",
                    staging_root=staging_root,
                    repo_root=repo_root,
                    config=config,
                    sanitize_connectors=args.sanitize_connectors,
                    redact_local_paths=args.redact_local_paths,
                )
            if args.include_downloaded_assets:
                extension = loader.env.get("ASSET_EXTENSION")
                asset = loader.latest_asset_path(extension)
                if asset:
                    stage_external_file(
                        rows,
                        source=asset,
                        zip_path=f"{PACKAGE_ROOT}/{spec.sheet_name}/downloads/{asset.name}",
                        test_id=spec.test_id,
                        suffix=spec.suffix,
                        category="asset",
                        staging_root=staging_root,
                        repo_root=repo_root,
                        config=config,
                        sanitize_connectors=False,
                        redact_local_paths=False,
                    )

        if args.excel and not minimal_specs:
            excel_path = resolve_repo_path(repo_root, args.excel)
            if excel_path.exists():
                stage_excel(rows, excel_path, staging_root, repo_root, config, args.sanitize_connectors, args.redact_local_paths)
            else:
                add_manifest_row(rows, test_id="_package", source_path=relative_to_repo(repo_root, excel_path), category="excel", included=False, exclusion_reason="missing")
                warnings.append(f"Excel not found: {excel_path}")

        readme_path = write_package_readme(staging_root, config, specs, args.sanitize_connectors)
        add_manifest_row(
            rows,
            test_id="_package",
            source_path="" if minimal_only else "generated",
            zip_path=f"{PACKAGE_ROOT}/README_PACKAGE.txt",
            category="readme",
            size_bytes=readme_path.stat().st_size,
            sha256=compute_sha256(readme_path),
            sanitized=args.sanitize_connectors,
            included=True,
        )
        # Manifest files are self-referential: their final content changes when
        # their own rows are written, so do not declare size/hash for themselves.
        for name in ("package_manifest.json", "package_manifest.csv"):
            add_manifest_row(
                rows,
                test_id="_package",
                source_path="" if minimal_only else "generated",
                zip_path=f"{PACKAGE_ROOT}/{name}",
                category="package_meta",
                size_bytes="",
                sha256="",
                sanitized=False,
                included=True,
                exclusion_reason="self-referential",
            )
        write_package_manifests(staging_root, rows)

        if args.dry_run:
            print_dry_run(rows)
            return 1 if args.strict and warnings else 0

        output = resolve_output_path(repo_root, args.output, args.export_dir, args.timestamp, args.timestamp_suffix)
        if not output:
            print("ERROR: --output or --export-dir is required unless --dry-run is used", file=sys.stderr)
            return 1
        if minimal_specs and any(spec.suffix in output.name for spec in minimal_specs):
            if output.exists():
                output.unlink()
            print(
                "ERROR: runtime T4 suffix appears in output filename",
                file=sys.stderr,
            )
            return 1
        create_zip(staging_root, output)
        audit_findings = audit_zip(output, args.sanitize_connectors)
        if minimal_only:
            audit_findings.extend(audit_minimal_publication_zip(output, minimal_specs[0]))
        for finding in audit_findings:
            print(f"WARN: audit: {finding}", file=sys.stderr)
        print(f"ZIP written: {relative_to_repo(repo_root, output)}")
        print(f"Included entries: {sum(1 for row in rows if row.get('included'))}")
        print(f"Excluded entries: {sum(1 for row in rows if not row.get('included'))}")
        if args.strict and (warnings or audit_findings):
            return 1
        return 0
    finally:
        shutil.rmtree(staging_parent, ignore_errors=True)


if __name__ == "__main__":
    raise SystemExit(main())
